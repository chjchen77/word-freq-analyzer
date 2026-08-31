#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
中文文本词频统计分析工具 v4.1
==============================
科研级中文文本词频统计。默认逐条保留原始记录，并从日期中识别年份和月份。
支持分类词典管理、正则/jieba 双模式、命中句子导出。

作者：陈浩杰
单位：澳门城市大学金融学院

v3.0 修复：
  - 修复子文件夹递归扫描遗漏 .xls 文件
  - 新增大文件分块读取（xlsx 流式读取，CSV >100MB 自动分块）
  - 支持逐条保留原始记录（默认）或按 公司×年份 聚合
  - Sheet2 改用 melt() 向量化构建，大数据集速度提升 10x+
  - 修复 pandas 2.x infer_datetime_format 废弃警告
  - 新增取消按钮，分析可随时中止
  - 新增 Excel 行数上限保护，超大结果自动截断并提示
  - 扫描后显示子文件夹统计，方便确认递归深度
  - 文件间自动回收内存，防止长时间运行 OOM
"""
from __future__ import annotations

import gc
import csv
import hashlib
import json
import os
import re
import sys
import threading
from datetime import date, datetime
# tkinter 仅 GUI 模式需要；服务器/CLI 模式下可无 tkinter 正常运行核心功能
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, simpledialog
    _HAS_TKINTER = True
except ImportError:
    _HAS_TKINTER = False
    # 无图形环境时提供占位基类，使 GUI 类定义可正常解析（不可实例化）
    # 用元类动态代理所有 tk.XXX / ttk.XXX 属性，避免逐一列举
    class _TkStub:  # noqa: E302
        def __init__(self, *a, **kw): pass
        def __init_subclass__(cls, **kw): pass
        def __getattr__(self, name): return self
        def __call__(self, *a, **kw): return self
        def __iter__(self): return iter([])

    class _MetaStub(type):
        """元类：将任意属性访问都返回 _TkStub，使 class Foo(tk.Bar) 可正常继承。"""
        def __getattr__(cls, name): return _TkStub

    class tk(metaclass=_MetaStub):  # type: ignore[no-redef]  # noqa: E302
        # tkinter 字符串常量（被 GUI 代码直接赋值使用）
        END = "end"; INSERT = "insert"; SEL = "sel"
        BOTH = "both"; LEFT = "left"; RIGHT = "right"; TOP = "top"; BOTTOM = "bottom"
        X = "x"; Y = "y"; W = "w"; E = "e"; N = "n"; S = "s"
        NW = "nw"; NE = "ne"; SW = "sw"; SE = "se"
        WORD = "word"; CHAR = "char"
        HORIZONTAL = "horizontal"; VERTICAL = "vertical"
        DISABLED = "disabled"; NORMAL = "normal"; ACTIVE = "active"
        CENTER = "center"; FLAT = "flat"; SUNKEN = "sunken"; RAISED = "raised"
        SINGLE = "single"; MULTIPLE = "multiple"; EXTENDED = "extended"
        GROOVE = "groove"; RIDGE = "ridge"; SOLID = "solid"

    class ttk(metaclass=_MetaStub):  # type: ignore[no-redef]  # noqa: E302
        pass

    class filedialog:  # type: ignore[no-redef]  # noqa: E302
        @staticmethod
        def askopenfilename(**kw): return ""
        @staticmethod
        def askopenfilenames(**kw): return ()
        @staticmethod
        def askdirectory(**kw): return ""
        @staticmethod
        def asksaveasfilename(**kw): return ""

    class messagebox:  # type: ignore[no-redef]  # noqa: E302
        @staticmethod
        def showinfo(*a, **kw): pass
        @staticmethod
        def showerror(*a, **kw): pass
        @staticmethod
        def showwarning(*a, **kw): pass
        @staticmethod
        def askyesno(*a, **kw): return False

    class simpledialog:  # type: ignore[no-redef]  # noqa: E302
        @staticmethod
        def askstring(*a, **kw): return None
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
# jieba 延迟导入：仅在实际使用分词时加载，避免启动时耗时

# ============================================================
# 常量
# ============================================================

# v3.1: 加入 .txt 支持（年报纯文本，文件名格式：代码_年份_公司名_标题_日期.txt）
SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv", ".txt"}

MAX_EXCEL_ROWS = 1_048_575  # Excel 行上限（减去表头）
BIG_CSV_THRESHOLD = 100 * 1024 * 1024  # 100MB
CHUNK_ROWS = 50_000
EXCEL_CHUNK_ROWS = 10_000

DATA_ENCODINGS = (
    "utf-8",
    "utf-8-sig",
    "gb18030",
    "gbk",
    "big5",
    "utf-16",
    "utf-16le",
    "utf-16be",
)

DEFAULT_STOPWORDS = frozenset({
    "的", "了", "在", "是", "我", "有", "和", "就", "不", "人",
    "都", "一", "一个", "上", "也", "很", "到", "说", "要", "去",
    "你", "会", "着", "没有", "看", "好", "自己", "这", "他", "她",
    "它", "们", "这个", "那个", "那", "什么", "如果", "但是", "因为",
    "所以", "可以", "已经", "而且", "或者", "虽然", "这些", "那些",
    "只是", "但", "而", "及", "与", "以", "为", "之", "其", "中",
    "对", "被", "从", "把", "将", "向", "个", "各", "等", "则",
    "能", "又", "该", "于", "当", "更", "还", "让", "用", "过",
    "后", "前", "下", "两", "所", "没", "这样", "那样", "怎么",
    "并", "给", "最", "些", "比", "做", "第", "如", "即", "且",
    "每", "应", "您", "哪", "得", "使", "才", "再", "还是", "及其",
    "以及", "因此", "由于", "其中", "可能", "已", "此", "需要",
    "通过", "进行", "没有", "不是", "如何", "或", "之间",
    " ", "\t", "\n", "\r",
    "，", "。", "、", "：", "；", "？", "！",
    """, """, "'", "'", "（", "）", "【", "】", "《", "》",
    "—", "…", "·",
    ",", ".", ":", ";", "?", "!", "(", ")", "[", "]",
    "{", "}", "-", "_", "/", "\\", "|", "@", "#", "%", "&", "*",
})

_NONE_LABEL = "（不选）"


def _excel_cell_value(value):
    """Normalize pandas/numpy scalars for row-wise xlsxwriter output."""
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "item"):
        try:
            value = value.item()
        except (TypeError, ValueError):
            pass
    if isinstance(value, str) and len(value) > 32767:
        return value[:32767]
    return value


def _normalize_column_names(values) -> list[str]:
    """把 Excel 原始表头规范成与 pandas 一致的、唯一的列名。"""
    columns: list[str] = []
    used: set[str] = set()
    next_suffix: dict[str, int] = {}
    for idx, value in enumerate(values):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            base = f"Unnamed: {idx}"
        else:
            base = str(value)
            if not base.strip():
                base = f"Unnamed: {idx}"

        candidate = base
        suffix = next_suffix.get(base, 0)
        while candidate in used:
            suffix += 1
            candidate = f"{base}.{suffix}"
        next_suffix[base] = suffix
        used.add(candidate)
        columns.append(candidate)
    return columns


def _detect_csv_separator(filepath: str, encoding: str) -> str:
    """从 CSV 首段探测常见分隔符；探测失败时保持逗号默认值。"""
    try:
        with open(filepath, "r", encoding=encoding, errors="strict", newline="") as fh:
            sample = fh.read(64 * 1024)
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except (OSError, UnicodeError, csv.Error):
        return ","


def _frame_from_excel_rows(rows: list[tuple], columns: list[str]) -> pd.DataFrame:
    """将 openpyxl 流式读取的行转成与 dtype=str 相近的 DataFrame。"""
    normalized_rows = []
    width = len(columns)
    for row in rows:
        values = list(row[:width])
        if len(values) < width:
            values.extend([None] * (width - len(values)))
        normalized_rows.append(values)
    if not normalized_rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(normalized_rows, columns=columns).astype("string")


def read_excel_chunked(filepath: str, chunksize: int = EXCEL_CHUNK_ROWS):
    """以 openpyxl read_only 模式分块读取 .xlsx 的第一张工作表。

    pandas.read_excel 默认读取第一张工作表，因此这里也固定使用第一张，
    避免“预览/分析因活动工作表不同而读到另一张表”的隐性不一致。
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("读取 .xlsx 文件需要安装 openpyxl 库：pip install openpyxl") from exc

    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0] if wb.worksheets else None
        if ws is None:
            return
        rows = ws.iter_rows(values_only=True)
        try:
            raw_header = next(rows)
        except StopIteration:
            return
        columns = _normalize_column_names(raw_header)
        buffer: list[tuple] = []
        for row in rows:
            buffer.append(row)
            if len(buffer) >= max(1, int(chunksize)):
                yield _frame_from_excel_rows(buffer, columns)
                buffer.clear()
        if buffer:
            yield _frame_from_excel_rows(buffer, columns)
    finally:
        wb.close()


def write_dataframes_xlsx(path: str, sheets: list[tuple[str, pd.DataFrame]]) -> None:
    """
    Write DataFrames row-by-row so xlsxwriter constant_memory mode remains safe.

    pandas.to_excel can lose all but the first column when xlsxwriter is in
    constant_memory mode because cells are not emitted strictly row by row.
    """
    try:
        import xlsxwriter
    except ImportError as exc:
        raise ValueError(
            "缺少 xlsxwriter 依赖，无法写出 Excel。请先执行：pip install xlsxwriter"
        ) from exc

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_name(f".{output_path.name}.{os.getpid()}.tmp")
    workbook = None
    try:
        workbook = xlsxwriter.Workbook(str(temp_path), {
            "constant_memory": True,
            "strings_to_urls": False,
        })
        for sheet_name, df in sheets:
            worksheet = workbook.add_worksheet(str(sheet_name)[:31])
            worksheet.write_row(0, 0, [str(c) for c in df.columns])
            for row_num, row in enumerate(df.itertuples(index=False, name=None), start=1):
                worksheet.write_row(row_num, 0, [_excel_cell_value(v) for v in row])
        workbook.close()
        workbook = None
        os.replace(temp_path, output_path)
    except Exception:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        try:
            temp_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise

# ============================================================
# 词典管理器
# ============================================================

class DictionaryManager:
    """管理 分类 → 关键词列表 映射。"""

    def __init__(self):
        self.data: dict[str, list[str]] = {}

    def categories(self) -> list[str]:
        return list(self.data.keys())

    def words(self, category: str) -> list[str]:
        return self.data.get(category, [])

    def add_category(self, name: str):
        if name and name not in self.data:
            self.data[name] = []

    def remove_category(self, name: str):
        self.data.pop(name, None)

    def add_word(self, category: str, word: str):
        word = word.strip()
        if category in self.data and word and word not in self.data[category]:
            self.data[category].append(word)

    def remove_word(self, category: str, word: str):
        if category in self.data:
            try:
                self.data[category].remove(word)
            except ValueError:
                pass

    def word_to_category(self) -> dict[str, str]:
        result = {}
        for cat, words in self.data.items():
            for w in words:
                result[w] = cat
        return result

    def find_duplicates(self) -> dict[str, list[str]]:
        """找出跨分类重复的关键词 → {词: [分类1, 分类2, ...]}。"""
        word_cats: dict[str, list[str]] = {}
        for cat, words in self.data.items():
            for w in words:
                word_cats.setdefault(w, []).append(cat)
        return {w: cats for w, cats in word_cats.items() if len(cats) > 1}

    def all_words(self) -> set[str]:
        return {w for words in self.data.values() for w in words}

    def total_word_count(self) -> int:
        return sum(len(ws) for ws in self.data.values())

    # ---- 导入 ----

    def import_file(self, path: str):
        """自动识别格式并导入词典（.xlsx / .xls / .txt）。"""
        ext = Path(path).suffix.lower()
        if ext in (".xlsx", ".xls"):
            self._import_excel(path)
        elif ext == ".txt":
            self._import_txt(path)
        else:
            raise ValueError(f"不支持的词典格式：{ext}（支持 .xlsx .xls .txt）")

    def _import_excel(self, path: str):
        df = pd.read_excel(path, header=None, dtype=str)
        if df.shape[1] < 2:
            raise ValueError("Excel 词典至少需要两列（分类、关键词）。")
        for row_idx, (_, row) in enumerate(df.iterrows()):
            cat = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            word = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            if row_idx == 0 and (
                cat.lower() in {"分类", "类别", "category", "cat", "tier_name"}
                and word.lower() in {"关键词", "关键字", "词语", "keyword", "term", "word"}
            ):
                continue
            if cat and word:
                self.add_category(cat)
                self.add_word(cat, word)

    def _import_txt(self, path: str):
        """格式：每行 分类：词1,词2,词3 或 分类:词1 词2 词3"""
        text = None
        for enc in ("utf-8", "utf-8-sig", "gb18030", "gbk"):
            try:
                with open(path, encoding=enc) as f:
                    text = f.read()
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError(f"无法读取文件（编码不支持）：{path}")

        for line in text.splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            sep = None
            for s in ("：", ":"):
                if s in line:
                    sep = s
                    break
            if sep is None:
                continue
            cat, words_part = line.split(sep, 1)
            cat = cat.strip()
            if not cat:
                continue
            self.add_category(cat)
            for w in re.split(r"[,，;；\s]+", words_part):
                w = w.strip()
                if w:
                    self.add_word(cat, w)

    def export_excel(self, path: str):
        rows = []
        for cat, words in self.data.items():
            for w in words:
                rows.append({"分类": cat, "关键词": w})
        pd.DataFrame(rows).to_excel(path, index=False)

    def clear(self):
        self.data.clear()

    def summary_text(self) -> str:
        parts = [f"  {cat}: {len(self.words(cat))} 个" for cat in self.categories()]
        return "\n".join(parts) if parts else "  （空）"


# ============================================================
# 工具函数
# ============================================================

def collect_data_files(folders: list[str]) -> tuple[list[str], dict[str, int], list[str]]:
    """
    递归扫描文件夹，返回 (文件列表, 子目录文件计数, 错误列表)。
    v3.0: followlinks=True 跟随符号链接；onerror 收集权限错误。
    v4.0: 用 realpath 去重防止符号链接循环导致无限递归。
    """
    files: list[str] = []
    dir_counts: dict[str, int] = {}
    errors: list[str] = []
    seen_realpaths: set[str] = set()  # 防止符号链接循环

    def _on_error(err):
        errors.append(str(err))

    for folder in folders:
        for root, dirs, filenames in os.walk(folder, followlinks=True, onerror=_on_error):
            real_root = os.path.realpath(root)
            if real_root in seen_realpaths:
                dirs[:] = []  # 已访问过，阻止 os.walk 继续深入
                continue
            seen_realpaths.add(real_root)
            found = 0
            for fname in sorted(filenames):
                if fname.startswith("~$") or fname.startswith("."):
                    continue
                if Path(fname).suffix.lower() in SUPPORTED_EXTENSIONS:
                    files.append(os.path.join(root, fname))
                    found += 1
            if found > 0:
                rel = os.path.relpath(root, folder)
                display = f"{os.path.basename(folder)}/{rel}" if rel != "." else os.path.basename(folder)
                dir_counts[display] = found
    return files, dir_counts, errors


def _parse_txt_filename(filepath: str) -> tuple[str, str]:
    """从 txt 文件名中解析公司代码和年份。
    支持格式：000001_2022_公司名_标题_日期.txt
    """
    path = Path(filepath)
    stem = path.stem
    code = ""
    year = ""
    code_match = re.search(r"(?<!\d)(\d{6})(?!\d)", stem)
    if code_match:
        code = code_match.group(1)
    else:
        parts = [p for p in re.split(r"[_\-\s]+", stem) if p]
        if parts:
            code = parts[0]
    year_match = re.search(r"((?:19|20)\d{2})", stem)
    if year_match:
        year = year_match.group(1)
    else:
        path_year_match = re.search(r"((?:19|20)\d{2})", str(path.parent))
        if path_year_match:
            year = path_year_match.group(1)
    return code, year


def read_data_file(filepath: str, nrows=None) -> pd.DataFrame:
    """读取单个数据文件（.xlsx / .xls / .csv / .txt），自动检测编码。"""
    ext = Path(filepath).suffix.lower()
    if ext == ".txt":
        code, year = _parse_txt_filename(filepath)
        for enc in ("utf-8", "utf-8-sig", "gb18030", "gbk", "big5"):
            try:
                with open(filepath, encoding=enc) as f:
                    content = f.read()
                return pd.DataFrame([{"公司代码": code, "年份": year, "文本内容": content}])
            except UnicodeDecodeError:
                continue
        raise ValueError(f"无法读取文件（编码不支持）：{filepath}")
    if ext == ".xlsx":
        chunks = read_excel_chunked(
            filepath,
            chunksize=max(1, int(nrows)) if nrows is not None else EXCEL_CHUNK_ROWS,
        )
        if nrows is not None:
            return next(chunks, pd.DataFrame())
        parts = list(chunks)
        return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()
    if ext == ".xls":
        try:
            return pd.read_excel(filepath, nrows=nrows, dtype=str)
        except ImportError:
            if ext == ".xls":
                raise ValueError(
                    f"读取 .xls 文件需要安装 xlrd 库：pip install xlrd\n"
                    f"文件：{filepath}"
                )
            raise
    # CSV
    for enc in DATA_ENCODINGS:
        try:
            sep = _detect_csv_separator(filepath, enc)
            return pd.read_csv(
                filepath, nrows=nrows, dtype=str, encoding=enc,
                sep=sep, low_memory=False,
            )
        except (UnicodeDecodeError, pd.errors.ParserError, pd.errors.EmptyDataError):
            continue
    raise ValueError(f"无法读取文件（编码不支持）：{filepath}")


def read_csv_chunked(filepath: str, chunksize: int = CHUNK_ROWS):
    """
    分块读取大 CSV 文件的生成器。
    v3.0: 先用小样本探测编码，再用确定的编码分块读取全文件。
    避免 generator 中途解码失败无法回退的问题。
    """
    # 第一步：探测正确的编码
    detected_enc = None
    for enc in DATA_ENCODINGS:
        try:
            sep = _detect_csv_separator(filepath, enc)
            pd.read_csv(filepath, dtype=str, encoding=enc, sep=sep, nrows=100)
            detected_enc = enc
            detected_sep = sep
            break
        except (UnicodeDecodeError, pd.errors.ParserError, pd.errors.EmptyDataError):
            continue
    if detected_enc is None:
        raise ValueError(f"无法读取文件（编码不支持）：{filepath}")

    # 第二步：用确定的编码分块读取
    reader = pd.read_csv(
        filepath, dtype=str, encoding=detected_enc,
        sep=detected_sep, low_memory=False, chunksize=chunksize,
    )
    for chunk in reader:
        yield chunk


def _read_columns_fast(filepath: str) -> list[str]:
    """
    快速读取文件列名（不加载数据）。
    v3.0: 对大 xlsx 使用 openpyxl read_only 模式，只读第一行；
    CSV 只读第一行。比 pd.read_excel(nrows=0) 快 10-100x。
    """
    ext = Path(filepath).suffix.lower()
    if ext == ".txt":
        return ["公司代码", "年份", "文本内容"]
    if ext in (".xlsx", ".xls"):
        # 优先用 openpyxl 快速读取第一行
        if ext == ".xlsx":
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.worksheets[0] if wb.worksheets else None
                for row in ws.iter_rows(max_row=1, values_only=True):
                    cols = _normalize_column_names(row)
                    wb.close()
                    return cols
                wb.close()
                return []
            except Exception:
                pass
        # 回退到 pandas（较慢但更兼容）
        try:
            df = pd.read_excel(filepath, nrows=0, dtype=str)
            return [str(c) for c in df.columns]
        except Exception:
            return []
    # CSV: 只读第一行
    for enc in DATA_ENCODINGS:
        try:
            sep = _detect_csv_separator(filepath, enc)
            df = pd.read_csv(filepath, nrows=0, dtype=str, encoding=enc, sep=sep)
            return [str(c) for c in df.columns]
        except (UnicodeDecodeError, pd.errors.ParserError, pd.errors.EmptyDataError):
            continue
    return []


def scan_all_columns(files: list[str]) -> tuple[list[str], dict[str, int]]:
    """
    扫描所有文件的列名（只读表头，不加载数据）。
    v3.0: 返回 (列名列表按出现频率降序, 列名→出现文件数)。
    频率高的列更可能是用户需要的公共列，优先展示。
    .xlsx 文件使用 _read_columns_fast() 的 openpyxl read_only 模式读取第一行，
    不因文件体积跳过表头扫描；否则单个大 Excel 会被误报为“发现 0 个数据列”。
    """
    col_freq: dict[str, int] = {}
    for f in files:
        try:
            ext = Path(f).suffix.lower()
            if ext == ".txt":
                for c in ["公司代码", "年份", "文本内容"]:
                    col_freq[c] = col_freq.get(c, 0) + 1
                continue
            columns = _read_columns_fast(f)
            for c in columns:
                col_freq[c] = col_freq.get(c, 0) + 1
        except Exception:
            continue
    # 按频率降序排列
    sorted_cols = sorted(col_freq.keys(), key=lambda x: -col_freq[x])
    return sorted_cols, col_freq


def fix_stock_code(code) -> str:
    """修复股票代码：剥离交易所后缀（.SH/.SZ/.BJ 等），数字补零至 6 位。"""
    if pd.isna(code):
        return ""
    s = str(code).strip()
    if not s:
        return ""
    # 剥离常见交易所后缀，如 600001.SH / 000001.SZ / 430001.BJ / 430001.NQ
    s = re.sub(r"\.(SH|SZ|BJ|NQ|HK|OQ)$", "", s, flags=re.IGNORECASE)
    try:
        return str(int(float(s))).zfill(6)
    except (ValueError, TypeError):
        return s


def parse_year_column(series: pd.Series) -> pd.Series:
    """智能解析年份列：支持数值、日期字符串、混合格式。

    修复：原实现用 30% 阈值选策略；若三种策略均低于 30%，整列返回 0，
    导致有效年份行被全量丢弃（静默数据损失）。
    新实现：按行合并多种解析策略，支持同一列内同时出现纯年份、日期、日期区间、
    “2009年”和 Excel 序列日期；无法解析的行置 0 交由调用方过滤。
    """
    if series.empty:
        return pd.Series(dtype=int)

    # 年报合理年份范围：1990–2030。
    # 1900 是 Excel 零值日期（NaT → 1900-01-01）的常见误解析结果；
    # 2099+ 通常来源于文件中的占位符日期，均应视为无效年份。
    _YEAR_MIN, _YEAR_MAX = 1990, 2030

    # 策略 1：直接数值
    numeric = pd.to_numeric(series, errors="coerce")
    valid_mask = numeric.between(_YEAR_MIN, _YEAR_MAX)
    direct_years = numeric.where(valid_mask, 0).fillna(0).astype(int)

    # 策略 2：日期解析。utc=True 保证混合格式/时区时仍返回可用的
    # Datetime Series，避免 .dt 在 object Series 上直接崩溃。
    date_strings = series.astype("string").str.strip()
    # 纯数字值优先按“年份/Excel 序列日期”解释，不送入日期解析，避免
    # 45292 被 pandas 当成 1970 年的纳秒时间戳。
    date_parse_input = date_strings.mask(numeric.notna())
    try:
        dates = pd.to_datetime(
            date_parse_input, errors="coerce", format="mixed", utc=True,
        )
    except (ValueError, TypeError):
        dates = pd.to_datetime(date_parse_input, errors="coerce", utc=True)
    years = dates.dt.year
    valid_year_mask = years.between(_YEAR_MIN, _YEAR_MAX)
    date_years = years.where(valid_year_mask, 0).fillna(0).astype(int)

    # 策略 3：Excel 序列日期（例如 45292 对应 2024 年）。
    excel_dates = pd.to_datetime(
        numeric, unit="D", origin="1899-12-30", errors="coerce", utc=True,
    )
    excel_years = excel_dates.dt.year
    excel_valid_mask = excel_years.between(_YEAR_MIN, _YEAR_MAX)
    excel_years = excel_years.where(excel_valid_mask, 0).fillna(0).astype(int)

    # 策略 4：正则提取 4 位年份，支持“2009年”、日期区间和多日期。
    extracted = date_strings.str.extract(r"((?:19|20)\d{2})", expand=False)
    extracted_num = pd.to_numeric(extracted, errors="coerce")
    valid_regex_mask = extracted_num.between(_YEAR_MIN, _YEAR_MAX)
    regex_years = extracted_num.where(valid_regex_mask, 0).fillna(0).astype(int)

    # 按行优先采用最直接的解释，再回退到日期、Excel 序列日期和正则提取。
    # 与“整列择一策略”相比，不会因少数混合格式行被错误丢弃。
    result = direct_years
    for candidate in (date_years, excel_years, regex_years):
        result = result.where(result > 0, candidate)
    return result.astype(int)


def parse_month_column(series: pd.Series) -> pd.Series:
    """从日期列提取月份；只有纯年份或无法确定月份时返回 0。

    支持普通日期、中文日期（2024年3月15日）、Excel 序列日期和日期区间。
    日期区间取其中第一个明确日期的月份，原始日期文本仍会完整写入结果。
    """
    if series.empty:
        return pd.Series(dtype=int)

    numeric = pd.to_numeric(series, errors="coerce")
    date_strings = series.astype("string").str.strip()
    normalized = (
        date_strings
        .str.replace("年", "-", regex=False)
        .str.replace("月", "-", regex=False)
        .str.replace("日", "", regex=False)
    )
    # “2024年”只表示年份，不能被 pandas 默认补成 2024-01；月份应保持未知。
    year_only_mask = normalized.str.fullmatch(r"(?:19|20)\d{2}-?")
    date_parse_input = normalized.mask(numeric.notna() | year_only_mask)
    try:
        dates = pd.to_datetime(
            date_parse_input, errors="coerce", format="mixed", utc=True,
        )
    except (ValueError, TypeError):
        dates = pd.to_datetime(date_parse_input, errors="coerce", utc=True)
    years = dates.dt.year
    date_months = dates.dt.month.where(
        years.between(1990, 2030), 0
    ).fillna(0).astype(int)

    # 仅把合理范围内的数字当作 Excel 序列日期，避免把“2024”误读成 1905 年。
    excel_numeric = numeric.where(numeric.between(20_000, 60_000))
    excel_dates = pd.to_datetime(
        excel_numeric, unit="D", origin="1899-12-30", errors="coerce", utc=True,
    )
    excel_months = excel_dates.dt.month.fillna(0).astype(int)

    # 日期解析失败时，从“2024年3月”“2024/3/15”等文本提取月份。
    extracted = normalized.str.extract(
        r"(?:(?:19|20)\d{2})\D{1,3}(\d{1,2})", expand=False
    )
    regex_months = pd.to_numeric(extracted, errors="coerce")
    regex_months = regex_months.where(regex_months.between(1, 12), 0).fillna(0).astype(int)

    result = date_months
    for candidate in (excel_months, regex_months):
        result = result.where(result > 0, candidate)
    return result.astype(int)


def _format_date_value(value) -> str:
    """将日期列值转换成适合写入结果表的显示文本，不丢失原始信息。"""
    if pd.isna(value):
        return ""
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return pd.Timestamp(value).strftime("%Y-%m-%d")
    text = str(value).strip()
    numeric = pd.to_numeric(text, errors="coerce")
    if pd.notna(numeric) and 20_000 <= numeric <= 60_000:
        try:
            return (pd.Timestamp("1899-12-30") + pd.to_timedelta(numeric, unit="D")) \
                .strftime("%Y-%m-%d")
        except (OverflowError, ValueError):
            pass
    return text


# 分句逻辑统一收敛到 sentence_split，避免与 count_mda_sentences 各存副本后漂移
# （分子=命中句、分母=MD&A句数，两者口径必须完全一致）。
from sentence_split import (  # noqa: E402
    MAX_SENT_LEN,
    is_table_like,
    split_sentences,
)


def _mp_worker_init(all_dict_words_: set, jieba_userdict_: str, use_regex_: bool):
    """ProcessPoolExecutor worker 初始化：非正则模式时在子进程中加载 jieba 词典。
    每个 worker 进程只执行一次，之后重用已初始化的 jieba 实例。
    """
    if not use_regex_ and all_dict_words_:
        import jieba as _jieba
        if jieba_userdict_ and os.path.isfile(jieba_userdict_):
            _jieba.load_userdict(jieba_userdict_)
        for w in all_dict_words_:
            _jieba.add_word(w)


def _process_one_file(
    fpath: str,
    col_stkcd: str,
    col_year: str,
    text_columns: list[str],
    keywords: list[str],
    use_regex: bool,
    all_dict_words: set[str],
    stopwords,
    use_stopwords: bool,
    export_sentences: bool,
    word_to_cat: dict[str, str],
    agg_rules: dict[str, str],
    preserve_rows: bool,
    cancel_event,
) -> tuple[list, list, str]:
    """处理单个文件，返回 (chunks, sents, log_msg)，异常会向上抛出。
    设计为独立函数（非闭包），可安全地在 ThreadPoolExecutor 中并发调用。
    """
    if cancel_event is not None and cancel_event.is_set():
        return [], [], "已取消"

    fname = os.path.basename(fpath)
    rel_dir = os.path.basename(os.path.dirname(fpath))
    display_name = f"{rel_dir}/{fname}" if rel_dir else fname

    ext = Path(fpath).suffix.lower()
    file_size = os.path.getsize(fpath)

    if ext == ".xlsx":
        file_chunks: list = []
        file_sents: list = []
        file_rows = file_hits = file_dropped = chunk_num = 0
        source_row_offset = 0
        for chunk_df in read_excel_chunked(fpath, EXCEL_CHUNK_ROWS):
            if cancel_event is not None and cancel_event.is_set():
                break
            chunk_num += 1
            result, rows, hits, sents, dropped = _process_chunk(
                chunk_df, col_stkcd, col_year, text_columns,
                keywords, use_regex, all_dict_words,
                stopwords, use_stopwords, export_sentences, word_to_cat,
                source_file=display_name,
                source_row_offset=source_row_offset,
                preserve_rows=preserve_rows,
            )
            source_row_offset += len(chunk_df)
            if result is not None:
                file_chunks.append(result)
                file_rows += rows
                file_hits += hits
                file_sents.extend(sents)
                file_dropped += dropped
        if file_chunks:
            if not preserve_rows:
                merged = pd.concat(file_chunks, ignore_index=True)
                merged = merged.groupby(["公司代码", "年份"], as_index=False).agg(agg_rules)
                file_chunks = [merged]
            dropped_note = f"（{file_dropped}行年份无效）" if file_dropped else ""
            return file_chunks, file_sents, f"{display_name}  {chunk_num}块 {file_rows}行 命中{file_hits}次{dropped_note}"
        return [], [], f"{display_name}  跳过：无有效数据"
    elif ext == ".csv" and file_size > BIG_CSV_THRESHOLD:
        file_chunks: list = []
        file_sents: list = []
        file_rows = file_hits = file_dropped = chunk_num = 0
        source_row_offset = 0
        for chunk_df in read_csv_chunked(fpath, CHUNK_ROWS):
            if cancel_event is not None and cancel_event.is_set():
                break
            chunk_num += 1
            result, rows, hits, sents, dropped = _process_chunk(
                chunk_df, col_stkcd, col_year, text_columns,
                keywords, use_regex, all_dict_words,
                stopwords, use_stopwords, export_sentences, word_to_cat,
                source_file=display_name,
                source_row_offset=source_row_offset,
                preserve_rows=preserve_rows,
            )
            source_row_offset += len(chunk_df)
            if result is not None:
                file_chunks.append(result)
                file_rows += rows
                file_hits += hits
                file_sents.extend(sents)
                file_dropped += dropped
        if file_chunks:
            if not preserve_rows:
                merged = pd.concat(file_chunks, ignore_index=True)
                merged = merged.groupby(["公司代码", "年份"], as_index=False).agg(agg_rules)
                file_chunks = [merged]
            dropped_note = f"（{file_dropped}行年份无效）" if file_dropped else ""
            return file_chunks, file_sents, f"{display_name}  {chunk_num}块 {file_rows}行 命中{file_hits}次{dropped_note}"
        return [], [], f"{display_name}  跳过：无有效数据"
    else:
        df = read_data_file(fpath)
        result, rows, hits, sents, dropped = _process_chunk(
            df, col_stkcd, col_year, text_columns,
            keywords, use_regex, all_dict_words,
            stopwords, use_stopwords, export_sentences, word_to_cat,
            source_file=display_name,
            source_row_offset=0,
            preserve_rows=preserve_rows,
        )
        del df
        if result is not None:
            dropped_note = f"（{dropped}行年份无效）" if dropped else ""
            return [result], sents, f"{display_name}  {rows}行 命中{hits}次{dropped_note}"
        return [], [], f"{display_name}  跳过：缺少必需列或无有效数据"


def load_stopwords_file(path: str) -> set[str]:
    words = set()
    for enc in ("utf-8", "utf-8-sig", "gb18030"):
        try:
            with open(path, encoding=enc) as f:
                for line in f:
                    w = line.strip()
                    if w:
                        words.add(w)
            return words
        except UnicodeDecodeError:
            continue
    return words


# ============================================================
# 面板数据分析引擎
# ============================================================

def _process_chunk(
    df: pd.DataFrame,
    col_stkcd: str,
    col_year: str,
    text_columns: list[str],
    keywords: list[str],
    use_regex: bool,
    all_dict_words: set[str],
    stopwords: set[str] | None,
    use_stopwords: bool,
    do_export_sentences: bool,
    word_to_cat: dict[str, str],
    *,
    source_file: str = "",
    source_row_offset: int = 0,
    preserve_rows: bool = False,
) -> tuple[pd.DataFrame | None, int, int, list[dict], int]:
    """
    处理单个 DataFrame（整文件或 CSV 块）。
    返回 (结果 DataFrame, 原始行数, 命中次数, 命中句子列表, 年份无效行数)。
    preserve_rows=True 时逐条保留原始记录；否则返回按公司×年份聚合的数据。
    """
    hit_sents: list[dict] = []

    if col_stkcd not in df.columns:
        return None, 0, 0, [], 0
    if col_year not in df.columns:
        return None, 0, 0, [], 0

    available_text = [c for c in text_columns if c in df.columns]
    if not available_text:
        return None, 0, 0, [], 0

    df = df.copy()
    raw_rows = len(df)  # v3.0: 在过滤前记录原始行数
    df["_source_file"] = source_file
    df["_source_row"] = range(source_row_offset + 2, source_row_offset + 2 + len(df))
    df["_stkcd"] = df[col_stkcd].apply(fix_stock_code)
    # 过滤空/无效公司代码，防止空字符串聚合成一条虚假"公司"记录污染面板
    _empty_stkcd_mask = df["_stkcd"].fillna("").str.strip() == ""
    if _empty_stkcd_mask.any():
        df = df[~_empty_stkcd_mask].copy()
    df["_year"] = parse_year_column(df[col_year])
    df["_month"] = parse_month_column(df[col_year])
    df["_date"] = df[col_year].map(_format_date_value)
    dropped_years = (df["_year"] == 0).sum()
    # 注意：_process_chunk 可在子进程中运行，log 不在作用域；年份警告由调用方汇总
    df = df[df["_year"] > 0].copy()  # v3.0: .copy() 防止 SettingWithCopyWarning
    if len(df) == 0:
        return None, raw_rows, 0, [], int(dropped_years)
    # 多列文本必须带分隔符拼接：直接 "".join 会让上一列结尾字与下一列开头字
    # 粘成不存在的词（如"生"+"态"→"生态"），既虚增词频也污染句子切分。
    text_series = df[available_text].fillna("").astype(str).agg("\n\n".join, axis=1)

    # 先切句，再在「切句结果」上计数，而非直接在原始文本上 str.count：
    #   · 原始文本被 PDF 折行切断的关键词（"生态\n环境"）无法匹配，实测漏计约 7%；
    #   · 分母 mda_sent 已剔除表格，分子若仍统计表格内命中则口径不一致。
    # 切句结果同时供下方句子导出复用，避免重复切分。
    sent_lists = [split_sentences(t) for t in text_series]
    analysis_series = pd.Series(["\n".join(s) for s in sent_lists], index=df.index)
    # 原始文本此后不再使用（计数走 analysis_series，导出走 sent_lists），
    # 提前释放为后续 190 个关键词的计数腾出余量。
    # 注意这并不降低峰值——峰值出现在上一行切句期间（原文与切句结果并存）。
    # 300 份 MD&A 实测增量 21 MB，按 CHUNK_ROWS=5 万行线性外推约 3.4 GB；
    # 真正削峰需改为分片切句。当前逐 txt 文件处理时每块仅 1 行，不触发该
    # 路径；若日后以 5 万行 CSV 作为输入，需先做分片。
    del text_series

    # 关键词匹配
    if use_regex:
        # 一次性构建所有关键词计数列，避免逐列 insert 导致 DataFrame 碎片化
        kw_counts = {
            kw: analysis_series.str.count(f"(?i){re.escape(kw)}")
            for kw in keywords
        }
        df = pd.concat([df, pd.DataFrame(kw_counts, index=df.index)], axis=1)
    else:
        import jieba  # 延迟导入：仅 jieba 模式才加载词典
        sw = stopwords if use_stopwords else None
        dict_set = all_dict_words

        # jieba 模式：列表累积后一次性构建 DataFrame（比 apply(pd.Series) 快 5-10x）
        # 同样在切句结果上分词，与正则模式保持一致口径
        rows_list = []
        for text in analysis_series:
            words = jieba.lcut(text)
            if use_stopwords and stopwords:
                words = [w for w in words if w not in stopwords]
            cnt: dict[str, int] = {}
            for w in words:
                if w in all_dict_words:
                    cnt[w] = cnt.get(w, 0) + 1
            rows_list.append({kw: cnt.get(kw, 0) for kw in keywords})
        kw_df = pd.DataFrame(rows_list, columns=keywords, index=df.index)
        for kw in keywords:
            df[kw] = kw_df[kw].values

    n_hits = int(df[keywords].sum().sum())

    # 提取命中句子
    # 用 str.__contains__（C 级 Boyer-Moore 搜索）替代 re.search：
    #   · 正确处理关键词互为子串的情况（不存在交替遮蔽问题）
    #   · 比 re.search 更快，且预计算 lower() 避免重复开销
    if do_export_sentences and n_hits > 0:
        match_mask = df[keywords].sum(axis=1) > 0
        row_pos = {idx: i for i, idx in enumerate(df.index)}
        for row_idx in df[match_mask].index:
            stkcd = df.loc[row_idx, "_stkcd"]
            year = int(df.loc[row_idx, "_year"])
            month = int(df.loc[row_idx, "_month"])
            row_kws = [kw for kw in keywords if df.loc[row_idx, kw] > 0]
            if not row_kws:
                continue
            # 每行预计算一次，避免在句子循环内重复 lower()
            kw_lower_pairs = [(kw, kw.lower()) for kw in row_kws]
            for sent in sent_lists[row_pos[row_idx]]:  # 复用上方切句结果
                sent_lower = sent.lower()
                for kw, kw_l in kw_lower_pairs:
                    if kw_l in sent_lower:
                        hit_sents.append({
                            "公司代码": stkcd,
                            "年份": year,
                            "月份": month if month > 0 else "",
                            "日期": df.loc[row_idx, "_date"],
                            "来源文件": df.loc[row_idx, "_source_file"],
                            "源文件行号": int(df.loc[row_idx, "_source_row"]),
                            "命中关键词": kw,
                            "分类": word_to_cat.get(kw, ""),
                            "命中句子": sent,
                        })

    if preserve_rows:
        # 逐条输出时保留日期、月份和源文件行号，确保相同公司同一天的多条记录也不合并。
        chunk = df[["_stkcd", "_year", "_month", "_date",
                    "_source_file", "_source_row"] + keywords].copy()
        chunk.columns = ["公司代码", "年份", "月份", "日期", "来源文件", "源文件行号"] + keywords
        chunk["月份"] = chunk["月份"].replace(0, pd.NA).astype("Int64")
        chunk["源文件行号"] = pd.to_numeric(chunk["源文件行号"], errors="coerce").astype("Int64")
        return chunk, raw_rows, n_hits, hit_sents, int(dropped_years)

    # 兼容旧的公司×年份汇总模式。
    chunk = df[["_stkcd", "_year"] + keywords].copy()
    chunk.columns = ["公司代码", "年份"] + keywords
    agg_rules = {kw: "sum" for kw in keywords}
    chunk_agg = chunk.groupby(["公司代码", "年份"], as_index=False).agg(agg_rules)

    return chunk_agg, raw_rows, n_hits, hit_sents, int(dropped_years)


def run_analysis(
    files: list[str],
    dict_mgr: DictionaryManager,
    col_stkcd: str,
    col_year: str,
    text_columns: list[str],
    output_path: str,
    *,
    use_regex: bool = True,
    use_stopwords: bool = False,
    stopwords: set[str] | None = None,
    use_tf: bool = False,
    export_sentences: bool = False,
    analyze_llm: bool = False,
    llm_api_key: str = "",
    llm_model: str = "qwen-plus",
    llm_base_url: str = "https://dashscope.aliyuncs.com/compatible-mode/v1",
    llm_max_sentences: int = 500,
    llm_max_workers: int = 4,
    llm_max_retries: int = 2,
    llm_cache_path: str | None = None,
    llm_system_prompt: str = "",
    analysis_workers: int = 1,
    preserve_rows: bool = True,
    jieba_userdict: str = "",
    progress_cb=None,
    log_cb=None,
    cancel_event: threading.Event | None = None,
):
    def log(msg):
        if log_cb:
            log_cb(msg)

    def is_cancelled():
        return cancel_event is not None and cancel_event.is_set()

    word_to_cat = dict_mgr.word_to_category()
    all_dict_words = dict_mgr.all_words()
    keywords = sorted(all_dict_words)
    categories = list(dict_mgr.categories())  # 快照，防止线程不安全

    if not all_dict_words:
        raise ValueError("词典为空，请先添加或导入词典。")

    if analyze_llm and not export_sentences:
        export_sentences = True
        log("提示：已自动启用导出命中句子，因为 LLM 句子分析依赖命中句子。")
    # 命中句子无人工上限：Excel 行数由写出时的 MAX_EXCEL_ROWS 保护兜底。
    # 之前设 20 万上限会导致大规模语料（>20 万命中）无法完整导出句子，已移除。
    sentence_cap = 0

    # 用户可能把输出路径选在数据目录内；再次运行时必须排除上一次产生的
    # 结果 Excel、完整句子 CSV、断点 CSV 和超大结果目录，避免把结果反读成输入。
    _output_stem = os.path.splitext(output_path)[0]
    _excluded_files = {
        os.path.abspath(output_path),
        os.path.abspath(_output_stem + "_sentences.xlsx"),
        os.path.abspath(_output_stem + "_sentences_full.csv"),
    }
    _excluded_dirs = {
        os.path.abspath(_output_stem + "_checkpoint"),
        os.path.abspath(_output_stem + "_csv"),
    }

    def _path_in_dir(path: str, directory: str) -> bool:
        try:
            return os.path.commonpath([os.path.abspath(path), directory]) == directory
        except ValueError:
            return False

    _input_files_before_filter = len(files)
    files = [
        f for f in files
        if os.path.abspath(f) not in _excluded_files
        and not any(_path_in_dir(f, d) for d in _excluded_dirs)
    ]
    if len(files) != _input_files_before_filter:
        log(f"已排除输出目录中的 {_input_files_before_filter - len(files)} 个旧结果文件，避免重复分析。")

    # v3.0: 跨分类重复关键词警告（同一词在多个分类中只会归入最后一个分类）
    duplicates = dict_mgr.find_duplicates()
    if duplicates:
        dup_items = [f"「{w}」→{cats}" for w, cats in list(duplicates.items())[:10]]
        log(f"警告：以下关键词在多个分类中重复出现（仅归入最后一个分类）：\n  " + "\n  ".join(dup_items))
        if len(duplicates) > 10:
            log(f"  …共 {len(duplicates)} 个重复词")
    # 构建词典诊断数据，写入输出文件
    dup_rows = [{"关键词": w, "所在分类（全部）": "、".join(cats), "实际归入分类": cats[-1],
                 "说明": "重复词：同一关键词出现在多个分类，仅计入最后一个分类"}
                for w, cats in (duplicates or {}).items()]

    all_diag_rows = dup_rows
    sheet_dict_diag = pd.DataFrame(all_diag_rows) if all_diag_rows else pd.DataFrame(
        columns=["关键词", "所在分类（全部）", "实际归入分类", "说明"]
    )

    # v3.0: 列名碰撞检查 — 关键词/分类名不能与内部列名冲突
    _reserved = {
        "公司代码", "年份", "月份", "日期", "来源文件", "源文件行号", "总计",
        "_stkcd", "_year", "_month", "_date",
    }
    bad_kws = [kw for kw in keywords if kw in _reserved]
    if bad_kws:
        log(f"警告：以下关键词与系统保留列名冲突，已自动跳过：{bad_kws}")
        keywords = [kw for kw in keywords if kw not in _reserved]
        all_dict_words -= _reserved
    bad_cats = [c for c in categories if c in _reserved]
    if bad_cats:
        log(f"警告：以下分类名与系统保留列名冲突，已自动跳过：{bad_cats}")
        categories = [c for c in categories if c not in _reserved]
        bad_cat_kws = [kw for kw in keywords if word_to_cat.get(kw) in bad_cats]
        if bad_cat_kws:
            log(f"警告：属于冲突分类的关键词也已跳过：{bad_cat_kws[:20]}")
            keywords = [kw for kw in keywords if kw not in bad_cat_kws]
            all_dict_words.difference_update(bad_cat_kws)
    valid_categories = set(categories)
    word_to_cat = {kw: cat for kw, cat in word_to_cat.items() if cat in valid_categories}
    # 过滤后若关键词全部被移除，给出明确的参数错误（而非后续聚合时的"无文件"误导）
    if not keywords:
        raise ValueError(
            "过滤与系统保留列名冲突的关键词后，词典已空。\n"
            f"被跳过的关键词：{bad_kws}\n"
            "请修改词典，避免使用「公司代码」「年份」「总计」等系统保留字作为关键词。"
        )

    if len(keywords) > 1000:
        log(f"提示：关键词数量较大（{len(keywords)}），分析可能较慢。")

    if use_regex:
        log("匹配模式：正则匹配（不区分大小写）")
    else:
        import jieba  # 延迟导入：仅 jieba 模式才加载词典
        if jieba_userdict and os.path.isfile(jieba_userdict):
            jieba.load_userdict(jieba_userdict)
            log(f"已加载 jieba 用户词典：{os.path.basename(jieba_userdict)}")
        for w in all_dict_words:
            jieba.add_word(w)
        log("匹配模式：jieba 分词")

    log(f"词典：{len(keywords)} 个关键词，{len(categories)} 个分类")
    if preserve_rows:
        log("统计方式：逐条保留原始记录，不按年份或月份汇总；结果包含日期和月份列。")
    else:
        log("统计方式：按公司代码×年份汇总。")
    log(f"公司代码列：{col_stkcd}，日期/年份列：{col_year}")
    log(f"文本列：{', '.join(text_columns)}")

    all_chunks: list[pd.DataFrame] = []
    hit_sentences: list[dict] = []
    total_files = len(files)
    agg_rules = {kw: "sum" for kw in keywords}
    _sent_lock = threading.Lock()
    _sent_capped = threading.Event()  # 并发模式下句子上限信号

    # ── 断点检查点（Checkpoint）──────────────────────────────────────────────
    # 每处理 _CKPT_INTERVAL 个文件写一次断点；进程崩溃后可从断点续跑，不丢数据。
    # 断点同时绑定输入文件、列配置、词典和匹配参数，避免用户改了设置后
    # 误复用旧结果。
    _CKPT_INTERVAL  = 5000
    _ckpt_dir       = os.path.splitext(output_path)[0] + "_checkpoint"
    _ckpt_chunks_path = os.path.join(_ckpt_dir, "chunks.csv")
    _ckpt_sents_path  = os.path.join(_ckpt_dir, "sentences.csv")
    _ckpt_done_path   = os.path.join(_ckpt_dir, "done_files.txt")
    _ckpt_meta_path   = os.path.join(_ckpt_dir, "metadata.json")

    _checkpoint_files = []
    for _f in sorted(files):
        try:
            _st = os.stat(_f)
            _checkpoint_files.append({
                "path": os.path.abspath(_f),
                "size": _st.st_size,
                "mtime_ns": _st.st_mtime_ns,
            })
        except OSError:
            _checkpoint_files.append({"path": os.path.abspath(_f), "missing": True})
    _checkpoint_payload = {
        "files": _checkpoint_files,
        "col_stkcd": col_stkcd,
        "col_year": col_year,
        "text_columns": list(text_columns),
        "preserve_rows": bool(preserve_rows),
        "dictionary": {cat: list(dict_mgr.words(cat)) for cat in categories},
        "keywords": list(keywords),
        "use_regex": bool(use_regex),
        "use_stopwords": bool(use_stopwords),
        "stopwords": sorted(stopwords or []),
        "use_tf": bool(use_tf),
        "export_sentences": bool(export_sentences),
        "analysis_workers": int(analysis_workers),
        "jieba_userdict": os.path.abspath(jieba_userdict) if jieba_userdict else "",
        "analyze_llm": bool(analyze_llm),
        "llm_model": llm_model,
        "llm_base_url": llm_base_url,
        "llm_max_sentences": int(llm_max_sentences),
        "llm_max_workers": int(llm_max_workers),
        "llm_max_retries": int(llm_max_retries),
        "llm_cache_path": os.path.abspath(llm_cache_path) if llm_cache_path else "",
        "llm_system_prompt": llm_system_prompt or "",
    }
    _checkpoint_signature = hashlib.sha1(
        json.dumps(_checkpoint_payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()

    _resume_done_set: set[str] = set()
    _checkpoint_usable = False
    try:
        if os.path.isfile(_ckpt_done_path):
            if os.path.isfile(_ckpt_meta_path):
                with open(_ckpt_meta_path, "r", encoding="utf-8") as _mf:
                    _saved_meta = json.load(_mf)
                _checkpoint_usable = _saved_meta.get("signature") == _checkpoint_signature
            if _checkpoint_usable:
                with open(_ckpt_done_path, "r", encoding="utf-8") as _ckf:
                    _resume_done_set = {ln.strip() for ln in _ckf if ln.strip()}
                log(f"⚡ 发现匹配当前配置的断点记录，已完成 {len(_resume_done_set)} 个文件，将从断点续跑。")
                if os.path.isfile(_ckpt_chunks_path):
                    _ck_df = pd.read_csv(_ckpt_chunks_path, dtype={"公司代码": str, "年份": str})
                    for _kc in [c for c in keywords if c in _ck_df.columns]:
                        _ck_df[_kc] = pd.to_numeric(_ck_df[_kc], errors="coerce").fillna(0).astype(int)
                    if preserve_rows and "月份" in _ck_df.columns:
                        _ck_df["月份"] = pd.to_numeric(_ck_df["月份"], errors="coerce").replace(0, pd.NA).astype("Int64")
                    if preserve_rows and "源文件行号" in _ck_df.columns:
                        _ck_df["源文件行号"] = pd.to_numeric(_ck_df["源文件行号"], errors="coerce").astype("Int64")
                    all_chunks.append(_ck_df)
                    log(f"  已加载断点 chunks：{len(_ck_df)} 行")
                    del _ck_df
                if os.path.isfile(_ckpt_sents_path):
                    _sk_df = pd.read_csv(_ckpt_sents_path)
                    hit_sentences.extend(_sk_df.to_dict("records"))
                    log(f"  已加载断点句子：{len(hit_sentences)} 条")
                    del _sk_df
                    if sentence_cap > 0 and len(hit_sentences) >= sentence_cap:
                        hit_sentences[:] = hit_sentences[:sentence_cap]
                files = [f for f in files if f not in _resume_done_set]
                log(f"  剩余待处理：{len(files)} 个文件（原始共 {total_files} 个）")
            else:
                log("⚠️  发现旧断点，但输入文件或分析配置已变化；旧断点已忽略，将从头分析。")
    except Exception as _ckpt_load_err:
        log(f"⚠️  断点加载失败，将从头开始：{_ckpt_load_err}")
        all_chunks.clear()
        hit_sentences.clear()
        _resume_done_set = set()

    # 元数据单独原子写入；即使程序随后中断，也不会产生“无指纹旧断点”。
    try:
        os.makedirs(_ckpt_dir, exist_ok=True)
        _tmp_meta = _ckpt_meta_path + ".tmp"
        with open(_tmp_meta, "w", encoding="utf-8") as _mf:
            json.dump({"signature": _checkpoint_signature, "payload": _checkpoint_payload},
                      _mf, ensure_ascii=False, indent=2, default=str)
        os.replace(_tmp_meta, _ckpt_meta_path)
    except Exception as _meta_err:
        log(f"⚠️  断点元数据保存失败（不影响分析）：{_meta_err}")

    _already_done: int = len(_resume_done_set)
    _ckpt_done_list: list[str] = list(_resume_done_set)

    def _save_checkpoint_now(label: str = "") -> None:
        """将当前进度（all_chunks / hit_sentences / 完成文件列表）原子写入断点目录。"""
        try:
            os.makedirs(_ckpt_dir, exist_ok=True)
            # chunks / sentences / done_files 都采用临时文件 + 原子替换，避免
            # 进程中断时出现“完成列表是新的、数据文件却只写了一半”的断点。
            if all_chunks:
                _tmp_c = pd.concat(all_chunks, ignore_index=True)
                _tmp_chunks_path = _ckpt_chunks_path + ".tmp"
                _tmp_c.to_csv(_tmp_chunks_path, index=False, encoding="utf-8-sig")
                os.replace(_tmp_chunks_path, _ckpt_chunks_path)
                del _tmp_c
            if hit_sentences:
                _tmp_sents_path = _ckpt_sents_path + ".tmp"
                pd.DataFrame(hit_sentences).to_csv(
                    _tmp_sents_path, index=False, encoding="utf-8-sig"
                )
                os.replace(_tmp_sents_path, _ckpt_sents_path)
            # 原子写 done_files（先写 .tmp 再 rename，防止写到一半崩溃）
            _tmp_done = _ckpt_done_path + ".tmp"
            with open(_tmp_done, "w", encoding="utf-8") as _f2:
                _f2.write("\n".join(_ckpt_done_list))
            os.replace(_tmp_done, _ckpt_done_path)
            _tag = f": {label}" if label else ""
            log(f"  [断点已保存{_tag}] 已完成 {len(_ckpt_done_list)}/{total_files} 文件")
        except Exception as _ce:
            log(f"  [断点保存失败，不影响分析] {_ce}")

    _workers = max(1, int(analysis_workers))

    if _workers == 1:
        # ── 单线程顺序模式（兼容低配设备，日志顺序清晰）────────────────
        do_export_sentences = export_sentences
        for idx, fpath in enumerate(files, _already_done + 1):
            if is_cancelled():
                log("用户已取消。")
                break
            fname = os.path.basename(fpath)
            rel_dir = os.path.basename(os.path.dirname(fpath))
            display_name = f"{rel_dir}/{fname}" if rel_dir else fname
            log(f"[{idx}/{total_files}] {display_name}")
            _file_succeeded = False
            try:
                chunks, sents, msg = _process_one_file(
                    fpath, col_stkcd, col_year, text_columns,
                    keywords, use_regex, all_dict_words,
                    stopwords, use_stopwords, do_export_sentences,
                    word_to_cat, agg_rules, preserve_rows, cancel_event,
                )
                log(f"  {msg.split('  ', 1)[-1]}" if "  " in msg else f"  {msg}")
                all_chunks.extend(chunks)
                hit_sentences.extend(sents)
                if sentence_cap > 0 and len(hit_sentences) > sentence_cap:
                    hit_sentences[:] = hit_sentences[:sentence_cap]
                    do_export_sentences = False
                    log("  命中句子已达上限，后续跳过提取。")
                _file_succeeded = True
            except Exception as e:
                log(f"  跳过（错误）：{e}")
            # 只有成功完成的文件才加入断点；错误文件必须允许下次重试。
            if _file_succeeded:
                _ckpt_done_list.append(fpath)
            if _file_succeeded and len(_ckpt_done_list) % _CKPT_INTERVAL == 0:
                _save_checkpoint_now(f"{len(_ckpt_done_list)}/{total_files}")
                # 压缩 all_chunks → 单 DataFrame，减少下次 concat 开销
                if len(all_chunks) > 1:
                    _compacted = pd.concat(all_chunks, ignore_index=True)
                    all_chunks.clear()
                    all_chunks.append(_compacted)
                    del _compacted
            gc.collect()
            if progress_cb:
                progress_cb(idx, total_files)

    else:
        # ── 多进程并发模式（真并行，充分利用多核 CPU）─────────────────
        log(f"并发模式：{_workers} 进程（多进程真并行，M4 Pro 等多核设备推荐）")
        from concurrent.futures import ProcessPoolExecutor, as_completed
        _jieba_userdict = jieba_userdict  # 传给 worker 初始化
        # 正则模式下子进程不需要 all_dict_words（只用 keywords 列表），
        # 避免将大型 set 序列化发送给每个 worker 进程
        _worker_dict_words = set() if use_regex else all_dict_words
        executor = ProcessPoolExecutor(
            max_workers=_workers,
            initializer=_mp_worker_init,
            initargs=(_worker_dict_words, _jieba_userdict, use_regex),
        )
        try:
            future_map = {
                executor.submit(
                    _process_one_file,
                    fpath, col_stkcd, col_year, text_columns,
                    keywords, use_regex, _worker_dict_words,
                    stopwords, use_stopwords, export_sentences,
                    word_to_cat, agg_rules, preserve_rows, None,  # 子进程不传 cancel_event
                ): (i + 1, fpath)
                for i, fpath in enumerate(files)
            }
            completed = 0
            for future in as_completed(future_map):
                idx, fpath = future_map[future]
                completed += 1
                _disp = completed + _already_done  # 显示时加上已跳过的断点文件数
                _file_succeeded = False
                try:
                    chunks, sents, msg = future.result()
                    log(f"[{_disp}/{total_files}] {msg}")
                    all_chunks.extend(chunks)
                    if sents and not _sent_capped.is_set():
                        with _sent_lock:
                            if not _sent_capped.is_set():
                                hit_sentences.extend(sents)
                                if sentence_cap > 0 and len(hit_sentences) >= sentence_cap:
                                    hit_sentences[:] = hit_sentences[:sentence_cap]
                                    _sent_capped.set()
                                    log("  命中句子已达上限，后续不再提取。")
                    _file_succeeded = True
                except Exception as e:
                    log(f"[{_disp}/{total_files}] 跳过（错误）：{e}")
                # 只有成功完成的 future 才能续跑跳过；异常文件下次重试。
                if _file_succeeded:
                    _ckpt_done_list.append(fpath)
                if _file_succeeded and len(_ckpt_done_list) % _CKPT_INTERVAL == 0:
                    _save_checkpoint_now(f"{len(_ckpt_done_list)}/{total_files}")
                    # 压缩 all_chunks，减少下次 concat 开销
                    if len(all_chunks) > 1:
                        _compacted = pd.concat(all_chunks, ignore_index=True)
                        all_chunks.clear()
                        all_chunks.append(_compacted)
                        del _compacted
                gc.collect()
                if progress_cb:
                    progress_cb(_disp, total_files)
                if is_cancelled():
                    executor.shutdown(wait=False, cancel_futures=True)
                    log("用户已取消，等待当前批次完成…")
                    break
        finally:
            executor.shutdown(wait=True)

    # 最终断点：全部文件处理完毕（或用户取消）后落盘一次，保证 100% 数据可恢复
    _save_checkpoint_now("全部文件处理完毕" if not is_cancelled() else "用户取消时保存")

    if is_cancelled():
        raise ValueError("分析已被用户取消。")

    # ---- 整理结果 ----
    if not all_chunks:
        raise ValueError("没有成功处理任何文件，请检查数据和列配置。")

    log("正在整理分析结果…")
    combined = pd.concat(all_chunks, ignore_index=True)
    del all_chunks
    gc.collect()

    if preserve_rows:
        # 新模式：不做任何公司/日期聚合，保留每一条有效原始记录。
        panel = combined
        id_cols = ["公司代码", "年份", "月份", "日期", "来源文件", "源文件行号"]
        main_sheet_name = "原始记录统计"
        keyword_sheet_name = "原始记录关键词"
        sort_cols = [c for c in id_cols if c in panel.columns]
        log(f"已保留 {len(panel):,} 条原始记录，未按年份或月份合并。")
    else:
        # 兼容旧模式：按公司×年份聚合。
        panel = combined.groupby(["公司代码", "年份"], as_index=False).agg(agg_rules)
        del combined
        gc.collect()
        id_cols = ["公司代码", "年份"]
        main_sheet_name = "公司年份分类统计"
        keyword_sheet_name = "关键词明细"
        sort_cols = id_cols

    panel.sort_values(sort_cols, inplace=True, kind="stable")
    panel.reset_index(drop=True, inplace=True)

    # v3.0: 提前构建分类→关键词映射快照（线程安全，不再依赖 dict_mgr）
    cat_to_kws: dict[str, list[str]] = {}
    for cat in categories:
        cat_to_kws[cat] = [kw for kw in keywords if word_to_cat.get(kw) == cat]

    # ---- Sheet1: 原始记录/公司×年份 分类统计 ----
    sheet1 = panel[id_cols].copy()
    for cat in categories:
        cat_kws = [kw for kw in cat_to_kws.get(cat, []) if kw in panel.columns]
        sheet1[cat] = panel[cat_kws].sum(axis=1).astype(int) if cat_kws else 0
    sheet1["总计"] = sheet1[categories].sum(axis=1)
    if use_tf:
        total_per_row = sheet1["总计"].replace(0, 1)  # 避免除零
        for cat in categories:
            sheet1[f"{cat}_占比"] = (sheet1[cat] / total_per_row).round(6)

    # ---- Sheet2: 记录/公司×年份×关键词 明细（v3.0: 向量化 melt）----
    log("正在构建关键词明细…")
    sheet2 = panel.melt(
        id_vars=id_cols,
        value_vars=keywords,
        var_name="关键词",
        value_name="次数",
    )
    sheet2["次数"] = sheet2["次数"].astype(int)
    sheet2 = sheet2[sheet2["次数"] > 0].copy()
    sheet2["分类"] = sheet2["关键词"].map(word_to_cat).fillna("")
    sheet2.sort_values(id_cols + ["分类", "关键词"], inplace=True, kind="stable")
    sheet2.reset_index(drop=True, inplace=True)

    # ---- Sheet3: 总体分类统计 ----
    cat_totals: dict[str, int] = {}
    for cat in categories:
        cat_kws = [kw for kw in cat_to_kws.get(cat, []) if kw in panel.columns]
        cat_totals[cat] = int(panel[cat_kws].sum().sum()) if cat_kws else 0
    grand_total = sum(cat_totals.values())
    rows_s3 = []
    for cat, total in sorted(cat_totals.items(), key=lambda x: -x[1]):
        rows_s3.append({
            "分类": cat,
            "总次数": total,
            "占比(%)": round(total / grand_total * 100, 2) if grand_total else 0,
        })
    sheet3 = pd.DataFrame(rows_s3)

    del panel
    gc.collect()

    # ---- 构建句子 DataFrame（尽早转换并释放 list 内存）----
    df_sent = None
    sheet4_rows = 0
    if hit_sentences:
        df_sent = pd.DataFrame(hit_sentences)
        del hit_sentences  # 立即释放原始 list，节省 1-2 GB
        gc.collect()
        _sort_cols = [c for c in ["公司代码", "年份", "分类", "命中关键词", "命中句子"]
                      if c in df_sent.columns]
        if _sort_cols:
            df_sent = df_sent.sort_values(_sort_cols, kind="stable").reset_index(drop=True)
        sheet4_rows = len(df_sent)

    # ---- Excel 行数上限保护 — 先保存完整 CSV，再截断写 Excel ----
    truncated = False
    if len(sheet1) > MAX_EXCEL_ROWS or len(sheet2) > MAX_EXCEL_ROWS:
        truncated = True
        csv_dir = os.path.splitext(output_path)[0] + "_csv"
        os.makedirs(csv_dir, exist_ok=True)
        log(f"数据量超过 Excel 行上限，正在先导出完整 CSV…")
        sheet1.to_csv(os.path.join(csv_dir, f"{main_sheet_name}.csv"), index=False, encoding="utf-8-sig")
        sheet2.to_csv(os.path.join(csv_dir, f"{keyword_sheet_name}.csv"), index=False, encoding="utf-8-sig")
        sheet3.to_csv(os.path.join(csv_dir, "分类汇总.csv"), index=False, encoding="utf-8-sig")
        log(f"完整数据已保存至：{csv_dir}")
    if len(sheet1) > MAX_EXCEL_ROWS:
        log(f"警告：Sheet1 共 {len(sheet1)} 行，Excel 上限 {MAX_EXCEL_ROWS}，Excel 中已截断。")
        sheet1 = sheet1.head(MAX_EXCEL_ROWS)
    if len(sheet2) > MAX_EXCEL_ROWS:
        log(f"警告：Sheet2 共 {len(sheet2)} 行，Excel 上限 {MAX_EXCEL_ROWS}，Excel 中已截断。")
        sheet2 = sheet2.head(MAX_EXCEL_ROWS)

    # ---- 先写词频结果 Excel（LLM 开始前落盘，防止 OOM 崩溃导致数据全丢）----
    log("正在写入词频结果 Excel…")
    _explain_rows = [
        {"项目": "统计方式", "说明": "当前默认逐条保留每一条有效原始记录，不按公司、年份或月份合并；每条记录均保留日期、月份、来源文件和源文件行号。" if preserve_rows else "当前按公司代码×年份汇总；同一公司同一年内的多条记录会合并。"},
        {"项目": "Sheet1 计数含义", "说明": "各分类列的数值为关键词在该条原始记录文本中的【出现次数】（含重复），不等于命中句子数。若同一句话中关键词出现3次，计3次。" if preserve_rows else "各分类列的数值为关键词在该公司-年份文本中的【出现次数】（含重复），不等于命中句子数。若同一句话中关键词出现3次，计3次。"},
        {"项目": "独立命中句子文件", "说明": "每条记录为一个关键词在一个句子中的命中实例。若同一句子命中同一关键词N次，主表计N次，但句子文件仅记录该句子一次。"},
        {"项目": "分类占比列（_占比）", "说明": "= 该分类出现次数 / 该行所有分类出现次数之和。这是分类构成比，不是传统TF词频（词频/文档总词数）。"},
        {"项目": "LLM分析维度", "说明": "LLM时间指向/语态/句子类型/确定性/量化属性/语气均基于关键词所在句子的语义判断，temperature=0确保可复现。"},
        {"项目": "重复关键词处理", "说明": "同一关键词出现在多个分类时，仅归入词典中最后定义的分类。详见【词典诊断】sheet。"},
        {"项目": "句子提取规则", "说明": "以。！？为句子边界切分；连续空行视为段落边界；单个换行视为PDF断行删除。最短句子长度=8字。"},
    ]
    # 第一步：只写主表格（不含命中句子）
    # 命中句子单独写入 _sentences.xlsx，避免大文件与大 DataFrame 同时驻留内存
    write_dataframes_xlsx(output_path, [
        (main_sheet_name, sheet1),
        (keyword_sheet_name, sheet2),
        ("分类汇总", sheet3),
        ("词典诊断", sheet_dict_diag),
        ("分析说明", pd.DataFrame(_explain_rows)),
    ])
    log(f"词频结果已保存至：{output_path}")

    # 词频 Excel 已落盘 → 断点目录使命完成，清理掉节省磁盘空间
    try:
        import shutil as _shutil
        if os.path.isdir(_ckpt_dir):
            _shutil.rmtree(_ckpt_dir)
            log("断点目录已清理。")
    except Exception as _ce:
        log(f"断点目录清理失败（不影响结果）：{_ce}")

    # ---- 立即释放主表格 DataFrame，为后续句子写入和 LLM 腾出内存 ----
    _n_sheet1 = len(sheet1)
    _n_companies = sheet1["公司代码"].nunique() if _n_sheet1 > 0 else 0
    _year_min = int(sheet1["年份"].min()) if _n_sheet1 > 0 else 0
    _year_max = int(sheet1["年份"].max()) if _n_sheet1 > 0 else 0
    _n_sheet2 = len(sheet2)
    _n_sheet3 = len(sheet3)
    del sheet1, sheet2, sheet3, sheet_dict_diag
    gc.collect()

    # 第二步：写命中句子（不含 LLM 列）到独立文件，LLM 完成后覆盖写入
    # 独立文件 = 永远不需要 openpyxl 加载大 xlsx，彻底避免 append OOM
    _sent_path = os.path.splitext(output_path)[0] + "_sentences.xlsx"
    _llm_status_dict = None  # 初始化，避免 df_sent 为 None 时末尾日志 NameError
    _llm_conf_dict = None
    if df_sent is not None:
        # 【关键保障】先把全量句子存成完整 CSV，防止 OOM 崩溃后数据丢失
        # 无论后续 LLM/Excel 写入是否成功，这份 CSV 始终完整保留
        _sent_full_csv = os.path.splitext(output_path)[0] + "_sentences_full.csv"
        log(f"正在保存全量命中句子 CSV（共 {len(df_sent)} 条）…")
        df_sent.to_csv(_sent_full_csv, index=False, encoding="utf-8-sig")
        log(f"全量句子已保存至：{_sent_full_csv}（即使后续崩溃数据也不丢失）")

        _n_pre = min(MAX_EXCEL_ROWS, len(df_sent))
        _df_pre = df_sent.head(_n_pre).copy()
        _df_pre.insert(0, "序号", range(1, _n_pre + 1))
        write_dataframes_xlsx(_sent_path, [("命中句子", _df_pre)])
        del _df_pre
        gc.collect()
        log(f"命中句子已保存至：{_sent_path}（LLM 分析完成后将更新）")

    # ---- LLM 句子分析（分批处理，每批 10 万条，避免 OOM）----
    _LLM_BATCH_SIZE = 10_000
    if df_sent is not None and analyze_llm:
        log("正在准备 LLM 句子分析...")
        if llm_cache_path is None:
            llm_cache_path = os.path.splitext(output_path)[0] + "_llm_cache.json"
        from llm_sentence_analyzer import LLMAnalyzerConfig, apply_llm_sentence_analysis
        llm_cfg = LLMAnalyzerConfig.from_inputs(
            api_key=llm_api_key,
            model=llm_model,
            base_url=llm_base_url,
            max_workers=llm_max_workers,
            max_sentences=llm_max_sentences,
            max_retries=llm_max_retries,
            cache_path=llm_cache_path,
            system_prompt=llm_system_prompt or None,
        )
        log(
            f"LLM 配置：模型 {llm_cfg.model}，唯一句子上限 {llm_cfg.max_sentences}，"
            f"缓存文件 {os.path.basename(llm_cfg.cache_path) if llm_cfg.cache_path else '未启用'}"
            + ("，使用自定义提示词" if llm_cfg.system_prompt else "")
        )
        n_total = len(df_sent)
        n_batches = max(1, (n_total + _LLM_BATCH_SIZE - 1) // _LLM_BATCH_SIZE)
        log(f"共 {n_total} 条句子，分 {n_batches} 批处理（每批最多 {_LLM_BATCH_SIZE} 条）")
        result_parts: list[pd.DataFrame] = []
        for _bi in range(n_batches):
            if cancel_event and cancel_event.is_set():
                break
            _s = _bi * _LLM_BATCH_SIZE
            _e = min(_s + _LLM_BATCH_SIZE, n_total)
            log(f"  第 {_bi + 1}/{n_batches} 批：行 {_s + 1}–{_e}")
            _batch = df_sent.iloc[_s:_e].copy()
            _batch_result = apply_llm_sentence_analysis(
                _batch, llm_cfg, log_cb=log, cancel_event=cancel_event,
            )
            result_parts.append(_batch_result)
            del _batch
            gc.collect()
        if result_parts:
            # 先释放旧 df_sent（无 LLM 列）再 concat，避免：旧+result_parts+新 三份同时驻留
            del df_sent
            gc.collect()
            df_sent = pd.concat(result_parts, ignore_index=True)
            del result_parts
            gc.collect()
        else:
            # 未完成任何批次（立即取消），保留原始 df_sent（无 LLM 列）
            del result_parts
            gc.collect()

    # 第三步：将带 LLM 列的命中句子覆盖写入独立文件（xlsxwriter constant_memory）
    # 绝不用 openpyxl mode='a'——那会把几百 MB 的 xlsx 全部加载进内存
    if df_sent is not None:
        # 超过 Excel 行上限时，先导出完整 CSV 防止数据丢失
        if len(df_sent) > MAX_EXCEL_ROWS:
            _sent_csv_path = os.path.splitext(output_path)[0] + "_sentences_full.csv"
            log(f"命中句子共 {len(df_sent)} 条，超过 Excel 行上限 {MAX_EXCEL_ROWS}，"
                f"正在导出完整 CSV 备份（流式写盘）…")
            df_sent.to_csv(_sent_csv_path, index=False, encoding="utf-8-sig")
            log(f"完整命中句子 CSV 已保存至：{_sent_csv_path}")
        _sent_out = df_sent
        if sentence_cap > 0 and len(_sent_out) > sentence_cap:
            _sent_out = _sent_out.head(sentence_cap).copy()
        elif len(_sent_out) > MAX_EXCEL_ROWS:
            _sent_out = _sent_out.head(MAX_EXCEL_ROWS).copy()
        else:
            _sent_out = _sent_out.copy()
        # 在 del df_sent 之前提取 LLM 统计（后面不能再访问 df_sent）
        _llm_status_dict = None
        _llm_conf_dict = None
        if analyze_llm and "LLM分析状态" in _sent_out.columns:
            _llm_status_dict = _sent_out["LLM分析状态"].value_counts(dropna=False).to_dict()
            if "LLM置信度" in _sent_out.columns:
                _llm_conf_dict = _sent_out[_sent_out["LLM分析状态"] == "成功"]["LLM置信度"].value_counts(dropna=False).to_dict()
        del df_sent  # 释放原始引用，GC 后内存降至仅 _sent_out
        gc.collect()
        _sent_out.insert(0, "序号", range(1, len(_sent_out) + 1))
        log("正在将命中句子写入 Excel…")
        write_dataframes_xlsx(_sent_path, [("命中句子", _sent_out)])
        log(f"命中句子已写入：{_sent_path}（{len(_sent_out)} 条）")
        del _sent_out
        gc.collect()

    _result_label = "原始记录" if preserve_rows else "面板记录"
    log(f"完成！{_n_sheet1} 条{_result_label}，{_n_companies} 家公司，年份 {_year_min}-{_year_max}")
    log(f"Sheet1: {main_sheet_name} ({_n_sheet1} 行)")
    log(f"Sheet2: {keyword_sheet_name} ({_n_sheet2} 行)")
    log(f"Sheet3: 分类汇总 ({_n_sheet3} 行)")
    if sheet4_rows:
        log(f"独立命中句子文件：{sheet4_rows} 条（详见 {os.path.basename(_sent_path)}）")
    if _llm_status_dict is not None:
        status_text = ", ".join(f"{k}={v}" for k, v in _llm_status_dict.items())
        log(f"LLM 句子分析状态：{status_text}")
        if _llm_conf_dict is not None:
            conf_text = ", ".join(f"{k}={v}" for k, v in _llm_conf_dict.items())
            log(f"LLM 置信度分布：{conf_text}")
    log(f"词频结果：{output_path}")
    log(f"命中句子：{_sent_path}")


# ============================================================
# LLM 续跑入口（词频已完成，只跑 LLM）
# ============================================================

def run_llm_analysis_only(
    sentences_source: str,
    output_path: str,
    *,
    llm_api_key: str = "",
    llm_model: str = "qwen-plus",
    llm_base_url: str = "https://dashscope.aliyuncs.com/compatible-mode/v1",
    llm_max_sentences: int = 0,
    llm_max_workers: int = 8,
    llm_max_retries: int = 2,
    llm_cache_path: str | None = None,
    llm_system_prompt: str = "",
    log_cb=None,
    cancel_event: threading.Event | None = None,
):
    """跳过词频阶段，直接对已有命中句子文件续跑 LLM 分析。

    典型用法：词频分析已完成（result.xlsx / result_sentences.xlsx 已落盘），
    但 LLM 因限流/OOM/网络中断而失败，无需重新跑 3-4 小时的词频，
    直接读取已有句子文件，结合 llm_cache.json（已分析的跳过），续跑剩余部分。

    sentences_source: 命中句子文件路径，优先传 _sentences_full.csv（完整数据），
                      次选 _sentences.xlsx（最多 MAX_EXCEL_ROWS 行）。
    output_path:      词频结果 xlsx 路径（用于推算 _sent_path 等输出路径）。
    """
    def log(msg):
        if log_cb:
            log_cb(msg)

    def is_cancelled():
        return cancel_event is not None and cancel_event.is_set()

    # ---- 读取命中句子 ----
    log(f"[LLM续跑] 正在读取命中句子：{sentences_source}…")
    src_path = Path(sentences_source)
    if not src_path.exists():
        raise FileNotFoundError(f"命中句子文件不存在：{sentences_source}")

    ext = src_path.suffix.lower()
    if ext == ".csv":
        df_sent = pd.read_csv(sentences_source, encoding="utf-8-sig", dtype=str)
    elif ext in (".xlsx", ".xls"):
        df_sent = pd.read_excel(sentences_source, engine="openpyxl", dtype=str)
    else:
        raise ValueError(f"不支持的格式：{ext}（仅支持 .csv / .xlsx / .xls）")

    # 年份列转回整数（CSV 读进来是字符串）
    if "年份" in df_sent.columns:
        df_sent["年份"] = pd.to_numeric(df_sent["年份"], errors="coerce").fillna(0).astype(int)

    # 去掉序号列（写出时重新生成），去掉旧 LLM 列（将被重新分析覆盖）
    _drop_cols = ["序号"] + [c for c in df_sent.columns if c.startswith("LLM")]
    df_sent = df_sent.drop(columns=[c for c in _drop_cols if c in df_sent.columns])

    log(f"[LLM续跑] 共读取 {len(df_sent)} 条命中句子。")
    gc.collect()

    # ---- LLM 配置 ----
    if llm_cache_path is None:
        llm_cache_path = os.path.splitext(output_path)[0] + "_llm_cache.json"
    from llm_sentence_analyzer import LLMAnalyzerConfig, apply_llm_sentence_analysis
    llm_cfg = LLMAnalyzerConfig.from_inputs(
        api_key=llm_api_key,
        model=llm_model,
        base_url=llm_base_url,
        max_workers=llm_max_workers,
        max_sentences=llm_max_sentences,
        max_retries=llm_max_retries,
        cache_path=llm_cache_path,
        system_prompt=llm_system_prompt or None,
    )
    limit_desc = "无上限" if llm_cfg.max_sentences <= 0 else str(llm_cfg.max_sentences)
    log(
        f"[LLM续跑] 模型 {llm_cfg.model}，上限 {limit_desc}，"
        f"线程 {llm_cfg.max_workers}，"
        f"缓存 {os.path.basename(llm_cfg.cache_path) if llm_cfg.cache_path else '未启用'}"
    )

    # ---- 分批 LLM 分析 ----
    _LLM_BATCH_SIZE = 10_000
    n_total = len(df_sent)
    n_batches = max(1, (n_total + _LLM_BATCH_SIZE - 1) // _LLM_BATCH_SIZE)
    log(f"[LLM续跑] 共 {n_total} 条句子，分 {n_batches} 批（每批最多 {_LLM_BATCH_SIZE} 条）")

    result_parts: list[pd.DataFrame] = []
    for _bi in range(n_batches):
        if is_cancelled():
            log("[LLM续跑] 用户已取消。")
            break
        _s = _bi * _LLM_BATCH_SIZE
        _e = min(_s + _LLM_BATCH_SIZE, n_total)
        log(f"  第 {_bi + 1}/{n_batches} 批：行 {_s + 1}–{_e}")
        _batch = df_sent.iloc[_s:_e].copy()
        _batch_result = apply_llm_sentence_analysis(
            _batch, llm_cfg, log_cb=log, cancel_event=cancel_event,
        )
        result_parts.append(_batch_result)
        del _batch
        gc.collect()

    if result_parts:
        del df_sent
        gc.collect()
        df_sent = pd.concat(result_parts, ignore_index=True)
        del result_parts
        gc.collect()
    else:
        del result_parts
        gc.collect()

    # ---- 写出结果 ----
    _sent_path = os.path.splitext(output_path)[0] + "_sentences.xlsx"

    # 超过 Excel 行上限时先导出完整 CSV，防止数据截断丢失
    if len(df_sent) > MAX_EXCEL_ROWS:
        _sent_csv_path = os.path.splitext(output_path)[0] + "_sentences_full.csv"
        log(f"[LLM续跑] 句子共 {len(df_sent)} 条，超 Excel 上限，导出完整 CSV…")
        df_sent.to_csv(_sent_csv_path, index=False, encoding="utf-8-sig")
        log(f"[LLM续跑] 完整 CSV 已保存至：{_sent_csv_path}")

    _sent_out = (df_sent.head(MAX_EXCEL_ROWS).copy()
                 if len(df_sent) > MAX_EXCEL_ROWS else df_sent.copy())
    del df_sent
    gc.collect()

    _sent_out.insert(0, "序号", range(1, len(_sent_out) + 1))
    log("[LLM续跑] 正在写入命中句子 Excel…")
    write_dataframes_xlsx(_sent_path, [("命中句子", _sent_out)])
    log(f"[LLM续跑] 完成！命中句子已写入：{_sent_path}（{len(_sent_out)} 条）")
    del _sent_out
    gc.collect()


# ============================================================
# 批量添加对话框
# ============================================================

class BatchAddDialog(tk.Toplevel):
    def __init__(self, parent, title="批量添加关键词"):
        super().__init__(parent)
        self.title(title)
        self.result: list[str] | None = None
        self.geometry("340x320")
        self.resizable(False, True)

        ttk.Label(self, text="每行输入一个关键词：").pack(padx=10, pady=(10, 0))
        self.text = tk.Text(self, width=36, height=14)
        self.text.pack(padx=10, pady=5, fill="both", expand=True)

        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=(0, 10))
        ttk.Button(btn_frame, text="确定", command=self._ok).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="取消", command=self.destroy).pack(side="left", padx=5)

        self.transient(parent)
        self.grab_set()
        self.text.focus_set()

    def _ok(self):
        raw = self.text.get("1.0", "end").strip()
        self.result = [w.strip() for w in raw.split("\n") if w.strip()]
        self.destroy()


# ============================================================
# 主界面
# ============================================================

class WordFreqApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("中文文本词频统计分析工具 v4.1 — 陈浩杰 | 澳门城市大学金融学院")
        self.geometry("1020x800")
        self.resizable(True, True)
        self.minsize(880, 680)

        # v3.0: 主题和样式
        self._setup_style()

        self.dict_mgr = DictionaryManager()
        self.folders: list[str] = []
        self.scanned_files: list[str] = []
        self.all_columns: list[str] = []
        self._col_freq: dict[str, int] = {}
        self._col_display_map: dict[str, str] = {}

        self.output_path = tk.StringVar()
        self.var_regex = tk.BooleanVar(value=True)
        self.var_stopwords = tk.BooleanVar(value=False)
        self.stopwords_path = tk.StringVar()
        self.var_tf = tk.BooleanVar(value=False)
        self.var_sentences = tk.BooleanVar(value=False)
        self.var_preserve_rows = tk.BooleanVar(value=True)
        self.analysis_workers_var = tk.StringVar(value="1")
        self.var_llm = tk.BooleanVar(value=False)
        self.llm_api_key_var = tk.StringVar(value=os.getenv("DASHSCOPE_API_KEY", ""))
        self.llm_show_key_var = tk.BooleanVar(value=False)
        self.llm_model_var = tk.StringVar(value=os.getenv("QWEN_MODEL", "qwen-plus"))
        self.llm_base_url_var = tk.StringVar(
            value=os.getenv("QWEN_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1")
        )
        self.llm_max_sentences_var = tk.StringVar(value="500")
        self.var_llm_no_limit = tk.BooleanVar(value=False)
        self._llm_sent_entry = None  # 句子上限 Entry 引用，用于启用/禁用
        self.llm_max_workers_var = tk.StringVar(value="4")
        self.llm_max_retries_var = tk.StringVar(value="2")
        self.llm_cache_custom_var = tk.BooleanVar(value=False)
        self.llm_cache_path_var = tk.StringVar(value="")
        self._llm_api_key_entry = None
        self._llm_cache_entry = None
        self._llm_cache_btn = None
        self._llm_prompt_text = None   # tk.Text 控件，保存自定义系统提示词
        self.jieba_dict_path = tk.StringVar()
        self._word_search_var = tk.StringVar()
        self._word_search_var.trace_add("write", self._on_word_search)
        self.current_file_var = tk.StringVar(value="就绪")

        # v3.0: 取消事件
        self._cancel_event = threading.Event()

        self._build_ui()

    # ================================================================
    #  样式配置
    # ================================================================

    def _setup_style(self):
        style = ttk.Style(self)
        available = style.theme_names()
        # macOS: aqua; Linux/Win: clam 或 vista
        for theme in ("aqua", "vista", "clam"):
            if theme in available:
                style.theme_use(theme)
                break

        # 标签页字体稍大
        style.configure("TNotebook.Tab", padding=(12, 6))
        # LabelFrame 标题加粗
        style.configure("TLabelframe.Label", font=("", 11, "bold"))
        # 按钮内边距
        style.configure("TButton", padding=(8, 4))
        # 大按钮
        style.configure("Accent.TButton", padding=(16, 8), font=("", 12, "bold"))

    # ================================================================
    #  UI 构建
    # ================================================================

    def _build_ui(self):
        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=10, pady=(10, 4))

        self._build_data_tab(notebook)
        self._build_dict_tab(notebook)
        self._build_settings_tab(notebook)
        self._build_run_tab(notebook)

        # v3.0: 底部状态栏
        status_frame = ttk.Frame(self, relief="sunken")
        status_frame.pack(fill="x", padx=10, pady=(0, 6))
        self._status_var = tk.StringVar(value="就绪  |  词典：0 词  |  数据：0 文件")
        ttk.Label(status_frame, textvariable=self._status_var,
                  font=("", 10)).pack(side="left", padx=8, pady=3)
        ttk.Label(status_frame, text="作者：陈浩杰 | 澳门城市大学金融学院",
                  font=("", 9), foreground="#666666").pack(side="right", padx=8, pady=3)

    # ---- 标签页1：数据选择 ----

    def _build_data_tab(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text=" 数据选择 ")

        # -- 文件夹 --
        frm_folder = ttk.LabelFrame(tab, text="文本文件夹（递归扫描所有子目录中的 .xlsx / .xls / .csv / .txt 文件）")
        frm_folder.pack(fill="x", padx=8, pady=(8, 4))

        top = ttk.Frame(frm_folder)
        top.pack(fill="x", padx=5, pady=5)
        self.folder_listbox = tk.Listbox(top, height=3)
        self.folder_listbox.pack(side="left", fill="both", expand=True)
        btns = ttk.Frame(top)
        btns.pack(side="right", padx=(5, 0))
        ttk.Button(btns, text="添加文件夹", command=self._add_folder).pack(fill="x", pady=1)
        ttk.Button(btns, text="移除选中", command=self._remove_folder).pack(fill="x", pady=1)
        ttk.Button(btns, text="扫描文件", command=self._scan_files).pack(fill="x", pady=1)

        # -- 文件列表 + 列配置 --
        mid = ttk.Frame(tab)
        mid.pack(fill="both", expand=False, padx=8, pady=4)
        mid.columnconfigure(0, weight=2)
        mid.columnconfigure(1, weight=1)
        mid.rowconfigure(0, weight=1)

        frm_files = ttk.LabelFrame(mid, text="扫描到的文件（点击预览数据）")
        frm_files.grid(row=0, column=0, sticky="nsew", padx=(0, 4))

        file_container = ttk.Frame(frm_files)
        file_container.pack(fill="both", expand=True, padx=5, pady=5)
        file_container.rowconfigure(0, weight=1)
        file_container.columnconfigure(0, weight=1)

        self.file_listbox = tk.Listbox(file_container, height=8)
        file_vsb = ttk.Scrollbar(file_container, orient="vertical", command=self.file_listbox.yview)
        file_hsb = ttk.Scrollbar(file_container, orient="horizontal", command=self.file_listbox.xview)
        self.file_listbox.configure(yscrollcommand=file_vsb.set, xscrollcommand=file_hsb.set)
        self.file_listbox.grid(row=0, column=0, sticky="nsew")
        file_vsb.grid(row=0, column=1, sticky="ns")
        file_hsb.grid(row=1, column=0, sticky="ew")
        self.file_listbox.bind("<<ListboxSelect>>", self._on_file_select)

        # 列配置面板
        frm_cols = ttk.LabelFrame(mid, text="列配置（面板数据必需）")
        frm_cols.grid(row=0, column=1, sticky="nsew")

        inner = ttk.Frame(frm_cols)
        inner.pack(fill="both", expand=True, padx=5, pady=5)

        ttk.Label(inner, text="公司代码列：").grid(row=0, column=0, sticky="w", pady=2)
        self.combo_stkcd = ttk.Combobox(inner, state="readonly", width=18)
        self.combo_stkcd.grid(row=0, column=1, sticky="ew", pady=2)

        ttk.Label(inner, text="日期/年份列：").grid(row=1, column=0, sticky="w", pady=2)
        self.combo_year = ttk.Combobox(inner, state="readonly", width=18)
        self.combo_year.grid(row=1, column=1, sticky="ew", pady=2)

        ttk.Label(inner, text="文本列（多选）：").grid(row=2, column=0, sticky="nw", pady=(4, 0))
        self.col_listbox = tk.Listbox(inner, height=4, selectmode="extended")
        col_sb = ttk.Scrollbar(inner, orient="vertical", command=self.col_listbox.yview)
        self.col_listbox.configure(yscrollcommand=col_sb.set)
        self.col_listbox.grid(row=2, column=1, sticky="nsew", pady=(4, 0))
        col_sb.grid(row=2, column=2, sticky="ns", pady=(4, 0))

        inner.columnconfigure(1, weight=1)
        inner.rowconfigure(2, weight=1)

        # -- 数据预览 --
        frm_preview = ttk.LabelFrame(tab, text="数据预览（前 100 行，点击上方文件自动加载）")
        frm_preview.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        container = ttk.Frame(frm_preview)
        container.pack(fill="both", expand=True, padx=5, pady=5)
        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)

        self.preview_tree = ttk.Treeview(container, show="headings")
        vsb = ttk.Scrollbar(container, orient="vertical", command=self.preview_tree.yview)
        hsb = ttk.Scrollbar(container, orient="horizontal", command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.preview_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

    # ---- 标签页2：词典管理 ----

    def _build_dict_tab(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text=" 词典管理 ")

        # ── 顶部工具栏 ──────────────────────────────────────────────
        toolbar = ttk.Frame(tab)
        toolbar.pack(fill="x", padx=8, pady=(8, 0))

        ttk.Button(toolbar, text="新建词典", command=self._dict_new).pack(side="left", padx=(0, 4))
        ttk.Button(toolbar, text="导入词典", command=self._dict_import).pack(side="left", padx=(0, 4))
        ttk.Button(toolbar, text="导出词典", command=self._dict_export).pack(side="left", padx=(0, 4))
        ttk.Separator(toolbar, orient="vertical").pack(side="left", fill="y", padx=6, pady=3)

        self.dict_info_var = tk.StringVar(value="当前词典：0 个分类，0 个关键词")
        ttk.Label(toolbar, textvariable=self.dict_info_var, foreground="#444444").pack(side="left")

        # ── 格式提示（折叠式）───────────────────────────────────────
        hint_frame = ttk.Frame(tab)
        hint_frame.pack(fill="x", padx=8, pady=(4, 0))
        hint_text = (
            "支持格式：① Excel (.xlsx/.xls)：第一列=分类，第二列=关键词  "
            "② 文本 (.txt)：每行 分类：词1,词2,词3"
        )
        ttk.Label(hint_frame, text=hint_text, foreground="#666666",
                  font=("", 9)).pack(anchor="w")

        ttk.Separator(tab, orient="horizontal").pack(fill="x", padx=8, pady=(6, 0))

        # ── 主面板（分类 左 + 关键词 右）────────────────────────────
        pane = ttk.Frame(tab)
        pane.pack(fill="both", expand=True, padx=8, pady=6)
        pane.columnconfigure(0, weight=1)
        pane.columnconfigure(1, weight=2)
        pane.rowconfigure(0, weight=1)

        # ── 左：分类列表 ─────────────────────────────────────────────
        frm_cat = ttk.LabelFrame(pane, text="分类")
        frm_cat.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        frm_cat.rowconfigure(0, weight=1)
        frm_cat.columnconfigure(0, weight=1)

        cat_container = ttk.Frame(frm_cat)
        cat_container.grid(row=0, column=0, sticky="nsew", padx=5, pady=(5, 0))
        cat_container.rowconfigure(0, weight=1)
        cat_container.columnconfigure(0, weight=1)

        self.cat_listbox = tk.Listbox(cat_container, height=12, activestyle="dotbox",
                                      selectbackground="#4a90d9", selectforeground="white")
        cat_vsb = ttk.Scrollbar(cat_container, orient="vertical", command=self.cat_listbox.yview)
        self.cat_listbox.configure(yscrollcommand=cat_vsb.set)
        self.cat_listbox.grid(row=0, column=0, sticky="nsew")
        cat_vsb.grid(row=0, column=1, sticky="ns")
        self.cat_listbox.bind("<<ListboxSelect>>", self._on_cat_select)

        cat_btns = ttk.Frame(frm_cat)
        cat_btns.grid(row=1, column=0, sticky="ew", padx=5, pady=5)
        ttk.Button(cat_btns, text="添加", width=6, command=self._cat_add).pack(side="left", padx=(0, 3))
        ttk.Button(cat_btns, text="重命名", width=7, command=self._cat_rename).pack(side="left", padx=(0, 3))
        ttk.Button(cat_btns, text="删除", width=6, command=self._cat_remove).pack(side="left")

        # ── 右：关键词列表 ───────────────────────────────────────────
        frm_words = ttk.LabelFrame(pane, text="关键词")
        frm_words.grid(row=0, column=1, sticky="nsew")
        frm_words.rowconfigure(1, weight=1)
        frm_words.columnconfigure(0, weight=1)

        # 搜索框
        search_row = ttk.Frame(frm_words)
        search_row.grid(row=0, column=0, sticky="ew", padx=5, pady=(5, 2))
        ttk.Label(search_row, text="搜索：").pack(side="left")
        self._word_search_entry = ttk.Entry(search_row, textvariable=self._word_search_var, width=20)
        self._word_search_entry.pack(side="left", fill="x", expand=True, padx=(2, 0))

        # 关键词列表（多选）
        word_container = ttk.Frame(frm_words)
        word_container.grid(row=1, column=0, sticky="nsew", padx=5, pady=0)
        word_container.rowconfigure(0, weight=1)
        word_container.columnconfigure(0, weight=1)

        self.word_listbox = tk.Listbox(word_container, height=10, selectmode="extended",
                                       activestyle="dotbox",
                                       selectbackground="#4a90d9", selectforeground="white")
        word_vsb = ttk.Scrollbar(word_container, orient="vertical", command=self.word_listbox.yview)
        self.word_listbox.configure(yscrollcommand=word_vsb.set)
        self.word_listbox.grid(row=0, column=0, sticky="nsew")
        word_vsb.grid(row=0, column=1, sticky="ns")

        # 关键词操作按钮
        word_btns = ttk.Frame(frm_words)
        word_btns.grid(row=2, column=0, sticky="ew", padx=5, pady=5)

        ttk.Button(word_btns, text="添加", width=6,
                   command=self._word_add).pack(side="left", padx=(0, 3))
        ttk.Button(word_btns, text="批量添加", width=8,
                   command=self._word_batch_add).pack(side="left", padx=(0, 3))
        ttk.Button(word_btns, text="删除所选", width=8,
                   command=self._word_remove).pack(side="left", padx=(0, 3))
        ttk.Button(word_btns, text="全选", width=5,
                   command=lambda: self.word_listbox.select_set(0, tk.END)).pack(side="left", padx=(0, 3))
        ttk.Button(word_btns, text="清空本类", width=8,
                   command=self._word_clear_all).pack(side="left")

    # ---- 标签页3：分析设置 ----

    def _build_settings_tab(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text=" 分析设置 ")
        pad = {"padx": 8, "pady": 4}

        # ── 匹配模式 ─────────────────────────────────────────────────
        frm_mode = ttk.LabelFrame(tab, text="匹配模式")
        frm_mode.pack(fill="x", **pad)
        ttk.Radiobutton(
            frm_mode, variable=self.var_regex, value=True,
            text="正则匹配（推荐）— re.findall 直接搜索，不区分大小写，支持中英文混合词库"
        ).pack(anchor="w", padx=8, pady=(4, 1))
        ttk.Radiobutton(
            frm_mode, variable=self.var_regex, value=False,
            text="jieba 分词匹配 — 先切词再匹配，适合纯中文精细分析（速度较慢）"
        ).pack(anchor="w", padx=8, pady=(1, 6))

        # ── 输出选项 ──────────────────────────────────────────────────
        frm_out_opts = ttk.LabelFrame(tab, text="输出选项")
        frm_out_opts.pack(fill="x", **pad)

        row_out0 = ttk.Frame(frm_out_opts)
        row_out0.pack(fill="x", padx=8, pady=(4, 2))
        ttk.Checkbutton(
            row_out0, variable=self.var_preserve_rows,
            text="逐条保留原始记录（推荐，不按年份或月份汇总；输出包含日期、月份、来源文件和行号）"
        ).pack(side="left")
        ttk.Label(
            frm_out_opts,
            text="取消勾选后恢复旧模式：按公司代码×年份汇总。",
            foreground="#666666",
        ).pack(anchor="w", padx=28, pady=(0, 2))

        row_out1 = ttk.Frame(frm_out_opts)
        row_out1.pack(fill="x", padx=8, pady=(4, 2))
        ttk.Checkbutton(
            row_out1, variable=self.var_tf,
            text="计算分类占比  （在 Sheet1 中为每个分类添加占比列 = 该分类次数 / 该行总命中次数）"
        ).pack(side="left")

        row_out2 = ttk.Frame(frm_out_opts)
        row_out2.pack(fill="x", padx=8, pady=(2, 4))
        ttk.Checkbutton(
            row_out2, variable=self.var_sentences,
            text="导出命中句子  （提取含关键词的原文句子到独立 _sentences.xlsx，用于论文引用验证，大数据集较慢）"
        ).pack(side="left")

        # ── 停用词 ───────────────────────────────────────────────────
        frm_stop = ttk.LabelFrame(tab, text="停用词过滤（仅 jieba 模式生效）")
        frm_stop.pack(fill="x", **pad)
        row_sw1 = ttk.Frame(frm_stop)
        row_sw1.pack(fill="x", padx=8, pady=(4, 2))
        ttk.Checkbutton(row_sw1, text="启用停用词过滤（内置 100+ 常用中文停用词）",
                        variable=self.var_stopwords).pack(side="left")
        row_sw2 = ttk.Frame(frm_stop)
        row_sw2.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Label(row_sw2, text="追加停用词文件：").pack(side="left")
        ttk.Entry(row_sw2, textvariable=self.stopwords_path, width=32,
                  state="readonly").pack(side="left", padx=(4, 6))
        ttk.Button(row_sw2, text="选择文件", command=self._choose_stopwords).pack(side="left")

        # ── 性能 & 输出路径 ──────────────────────────────────────────
        frm_misc = ttk.LabelFrame(tab, text="性能与输出")
        frm_misc.pack(fill="x", **pad)

        row_m1 = ttk.Frame(frm_misc)
        row_m1.pack(fill="x", padx=8, pady=(6, 3))
        ttk.Label(row_m1, text="并发文件数：").pack(side="left")
        ttk.Entry(row_m1, textvariable=self.analysis_workers_var, width=5).pack(side="left", padx=(4, 8))
        ttk.Label(row_m1, text="1 = 单线程  |  4–8 = 推荐（M4 Pro）  |  建议 ≤ 16",
                  foreground="#666666").pack(side="left")

        row_m2 = ttk.Frame(frm_misc)
        row_m2.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Label(row_m2, text="输出文件路径：").pack(side="left")
        ttk.Entry(row_m2, textvariable=self.output_path, width=42,
                  state="readonly").pack(side="left", padx=(4, 6), fill="x", expand=True)
        ttk.Button(row_m2, text="选择路径", command=self._choose_output).pack(side="left")

        # ── LLM 句子分析 ─────────────────────────────────────────────
        frm_llm = ttk.LabelFrame(tab, text="大语言模型句子分析（OpenAI 兼容接口）")
        frm_llm.pack(fill="x", **pad)

        row_l0 = ttk.Frame(frm_llm)
        row_l0.pack(fill="x", padx=8, pady=(6, 4))
        ttk.Checkbutton(
            row_l0, variable=self.var_llm,
            text="启用 LLM 分析  （对命中句子自动标注：时间指向 / 语态 / 句子类型 / 确定性 / 量化属性 / 语气）",
        ).pack(side="left")

        ttk.Separator(frm_llm, orient="horizontal").pack(fill="x", padx=8, pady=2)

        # Grid 布局让左侧标签右对齐，右侧控件填满
        grid_frm = ttk.Frame(frm_llm)
        grid_frm.pack(fill="x", padx=8, pady=4)
        grid_frm.columnconfigure(1, weight=1)

        # Row 0: API Key
        ttk.Label(grid_frm, text="API Key：", anchor="e").grid(
            row=0, column=0, sticky="e", padx=(0, 4), pady=3)
        key_row = ttk.Frame(grid_frm)
        key_row.grid(row=0, column=1, sticky="ew", pady=3)
        self._llm_api_key_entry = ttk.Entry(
            key_row, textvariable=self.llm_api_key_var, show="*")
        self._llm_api_key_entry.pack(side="left", fill="x", expand=True)
        ttk.Checkbutton(
            key_row, text="显示",
            variable=self.llm_show_key_var,
            command=self._toggle_api_key_visibility,
        ).pack(side="left", padx=(6, 0))

        # Row 1: Base URL
        ttk.Label(grid_frm, text="Base URL：", anchor="e").grid(
            row=1, column=0, sticky="e", padx=(0, 4), pady=3)
        url_row = ttk.Frame(grid_frm)
        url_row.grid(row=1, column=1, sticky="ew", pady=3)
        ttk.Entry(url_row, textvariable=self.llm_base_url_var).pack(
            side="left", fill="x", expand=True)
        ttk.Label(url_row, text=" 兼容 OpenAI 接口的任何服务",
                  foreground="#666666").pack(side="left")

        # Row 2: Model + 测试连接
        ttk.Label(grid_frm, text="模型名称：", anchor="e").grid(
            row=2, column=0, sticky="e", padx=(0, 4), pady=3)
        model_row = ttk.Frame(grid_frm)
        model_row.grid(row=2, column=1, sticky="ew", pady=3)
        ttk.Entry(model_row, textvariable=self.llm_model_var, width=24).pack(
            side="left", padx=(0, 8))
        ttk.Label(model_row, text="如：qwen3.7-max / qwen-plus / qwen-max / gpt-4o-mini",
                  foreground="#666666").pack(side="left")
        ttk.Button(model_row, text="测试连接",
                   command=self._test_llm_connection).pack(side="right")

        # Row 3: 句子上限 / 并发 / 重试
        ttk.Label(grid_frm, text="参数：", anchor="e").grid(
            row=3, column=0, sticky="e", padx=(0, 4), pady=3)
        param_row = ttk.Frame(grid_frm)
        param_row.grid(row=3, column=1, sticky="ew", pady=3)
        ttk.Label(param_row, text="句子上限").pack(side="left")
        self._llm_sent_entry = ttk.Entry(
            param_row, textvariable=self.llm_max_sentences_var, width=7)
        self._llm_sent_entry.pack(side="left", padx=(3, 4))
        ttk.Checkbutton(
            param_row, text="不限制",
            variable=self.var_llm_no_limit,
            command=self._on_llm_no_limit_toggle,
        ).pack(side="left", padx=(0, 14))
        ttk.Label(param_row, text="并发线程").pack(side="left")
        ttk.Entry(param_row, textvariable=self.llm_max_workers_var, width=5).pack(
            side="left", padx=(3, 14))
        ttk.Label(param_row, text="最大重试").pack(side="left")
        ttk.Entry(param_row, textvariable=self.llm_max_retries_var, width=5).pack(
            side="left", padx=(3, 0))
        ttk.Label(param_row,
                  text="  每句约 200–500 token，正式跑前请估算费用",
                  foreground="#cc6600").pack(side="left")

        # Row 4: 缓存文件
        ttk.Label(grid_frm, text="缓存文件：", anchor="e").grid(
            row=4, column=0, sticky="e", padx=(0, 4), pady=3)
        cache_row = ttk.Frame(grid_frm)
        cache_row.grid(row=4, column=1, sticky="ew", pady=(3, 8))
        ttk.Checkbutton(
            cache_row, text="自定义路径",
            variable=self.llm_cache_custom_var,
            command=self._on_llm_cache_toggle,
        ).pack(side="left", padx=(0, 4))
        self._llm_cache_entry = ttk.Entry(
            cache_row, textvariable=self.llm_cache_path_var, state="disabled")
        self._llm_cache_entry.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self._llm_cache_btn = ttk.Button(
            cache_row, text="选择", state="disabled",
            command=self._choose_llm_cache_path,
        )
        self._llm_cache_btn.pack(side="left", padx=(0, 8))
        ttk.Label(cache_row,
                  text="不勾选则与输出文件同目录，重复运行自动跳过已分析句子",
                  foreground="#666666").pack(side="left")

        # ── 自定义系统提示词 ──────────────────────────────────────────
        frm_prompt = ttk.LabelFrame(frm_llm, text="自定义系统提示词（可选，留空使用内置模板）")
        frm_prompt.pack(fill="x", padx=8, pady=(0, 8))

        # 顶部操作栏
        prompt_toolbar = ttk.Frame(frm_prompt)
        prompt_toolbar.pack(fill="x", padx=6, pady=(4, 0))
        ttk.Label(
            prompt_toolbar,
            text="修改后将替换内置分析维度说明（格式要求见内置模板）；不修改则留空即可。",
            foreground="#666666",
        ).pack(side="left")
        ttk.Button(
            prompt_toolbar,
            text="重置为默认",
            command=self._reset_llm_prompt,
        ).pack(side="right", padx=(4, 0))

        # 可滚动文本框
        from tkinter import scrolledtext as _st
        self._llm_prompt_text = _st.ScrolledText(
            frm_prompt, height=10, wrap="word", font=("Courier", 10),
            undo=True,
        )
        self._llm_prompt_text.pack(fill="x", padx=6, pady=(4, 6))
        # 默认留空（留空 = 使用内置提示词）；用户点"重置为默认"可查看/恢复模板
        self._llm_prompt_text.insert("1.0", "")

        # ── jieba 自定义词典 ──────────────────────────────────────────
        frm_jieba = ttk.LabelFrame(tab, text="jieba 用户词典（仅 jieba 模式，可选）")
        frm_jieba.pack(fill="x", **pad)
        row_jb = ttk.Frame(frm_jieba)
        row_jb.pack(fill="x", padx=8, pady=6)
        ttk.Entry(row_jb, textvariable=self.jieba_dict_path, width=50,
                  state="readonly").pack(side="left", fill="x", expand=True, padx=(0, 6))
        ttk.Button(row_jb, text="选择文件", command=self._choose_jieba_dict).pack(side="left")

    # ---- 标签页4：运行分析 ----

    def _build_run_tab(self, notebook):
        tab = ttk.Frame(notebook)
        notebook.add(tab, text=" 运行分析 ")
        pad = {"padx": 8, "pady": 4}

        # v3.0: 开始 + 取消按钮
        btn_row = ttk.Frame(tab)
        btn_row.pack(pady=(15, 5))
        self.btn_start = ttk.Button(btn_row, text="  开始统计  ", command=self._start_analysis)
        self.btn_start.pack(side="left", padx=5)
        self.btn_cancel = ttk.Button(btn_row, text="  取消  ", command=self._cancel_analysis, state="disabled")
        self.btn_cancel.pack(side="left", padx=5)

        frm_prog = ttk.LabelFrame(tab, text="进度")
        frm_prog.pack(fill="x", **pad)
        self.progress = ttk.Progressbar(frm_prog, orient="horizontal", mode="determinate")
        self.progress.pack(fill="x", padx=5, pady=(5, 2))
        ttk.Label(frm_prog, textvariable=self.current_file_var).pack(
            anchor="w", padx=5, pady=(0, 5)
        )

        frm_log = ttk.LabelFrame(tab, text="运行日志")
        frm_log.pack(fill="both", expand=True, **pad)
        self.log_text = tk.Text(frm_log, wrap="word", height=18, state="disabled",
                               font=("Menlo", 11) if sys.platform == "darwin" else ("Consolas", 10))
        log_sb = ttk.Scrollbar(frm_log, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_sb.set)
        self.log_text.pack(side="left", fill="both", expand=True, padx=(5, 0), pady=5)
        log_sb.pack(side="right", fill="y", padx=(0, 5), pady=5)

        # v3.0: 日志颜色标签
        self.log_text.tag_configure("error", foreground="#cc3333")
        self.log_text.tag_configure("success", foreground="#228B22")
        self.log_text.tag_configure("warn", foreground="#cc8800")

    # ================================================================
    #  数据选择 — 事件
    # ================================================================

    def _add_folder(self):
        folder = filedialog.askdirectory(title="选择文本文件夹")
        if folder and folder not in self.folders:
            self.folders.append(folder)
            self.folder_listbox.insert(tk.END, folder)

    def _remove_folder(self):
        sel = self.folder_listbox.curselection()
        if sel:
            idx = sel[0]
            self.folder_listbox.delete(idx)
            self.folders.pop(idx)

    def _scan_files(self):
        if not self.folders:
            messagebox.showwarning("提示", "请先添加文件夹。")
            return

        # v3.0: 使用增强的 collect_data_files
        self.scanned_files, dir_counts, errors = collect_data_files(self.folders)
        self.file_listbox.delete(0, tk.END)
        for f in self.scanned_files:
            self.file_listbox.insert(tk.END, f)

        if not self.scanned_files:
            err_msg = ""
            if errors:
                err_msg = "\n\n扫描错误：\n" + "\n".join(errors[:5])
            messagebox.showinfo("提示", f"未找到 .xlsx / .xls / .csv / .txt 文件。{err_msg}")
            return

        # v3.0: 后台线程扫描列名（大 xlsx 读取慢，避免冻结 UI）
        self._status_var.set(f"正在扫描 {len(self.scanned_files)} 个文件的列名…")
        self.update_idletasks()

        scan_dir_counts = dir_counts
        scan_errors = errors

        def _do_scan():
            cols, freq = scan_all_columns(self.scanned_files)
            self.after(0, lambda: self._finish_scan(cols, freq, scan_dir_counts, scan_errors))

        threading.Thread(target=_do_scan, daemon=True).start()

    def _finish_scan(self, cols, freq, dir_counts, errors):
        """列名扫描完成后在主线程更新 UI。"""
        self.all_columns = cols
        self._col_freq = freq
        n_files = len(self.scanned_files)

        col_values = list(self.all_columns)  # 已按频率降序排列
        self.combo_stkcd["values"] = col_values
        self.combo_year["values"] = col_values
        self.col_listbox.delete(0, tk.END)
        self._col_display_map = {}
        for c in col_values:
            f = self._col_freq.get(c, 0)
            # 在列表中显示列名及其出现频率，方便用户判断
            label = f"{c}  [{f}/{n_files}]" if f < n_files else c
            self.col_listbox.insert(tk.END, label)
            self._col_display_map[label] = c

        # v3.0: 按频率加权的自动列匹配
        stkcd_patterns = (
            "股票代码", "证券代码", "公司代码", "companyid", "company_id",
            "企业id", "企业编号", "stkcd", "stock_code", "scode", "code",
        )
        year_patterns = ("年份", "year", "年", "日期", "时间", "date", "qtm")

        def _best_match(patterns):
            candidates = []
            for c in col_values:
                cl = c.lower()
                if any(k in cl for k in patterns):
                    candidates.append((c, self._col_freq.get(c, 0)))
            if candidates:
                candidates.sort(key=lambda x: -x[1])
                return candidates[0][0]
            return None

        best_stkcd = _best_match(stkcd_patterns)
        if best_stkcd:
            self.combo_stkcd.set(best_stkcd)

        best_year = _best_match(year_patterns)
        if best_year:
            self.combo_year.set(best_year)

        # v3.0: 文本列自动选择（匹配常见文本列名，高频优先）
        text_patterns = ("reply", "回复", "text", "文本", "内容", "content", "qsubj",
                         "问题", "question", "answer", "描述", "说明", "摘要", "正文",
                         "subject", "title", "标题", "body")
        auto_text_indices = []
        for idx, label in enumerate(list(self._col_display_map.keys())):
            real_col = self._col_display_map[label]
            cl = real_col.lower()
            if any(k in cl for k in text_patterns):
                auto_text_indices.append(idx)
        if auto_text_indices:
            for i in auto_text_indices:
                self.col_listbox.selection_set(i)

        # v3.0: 自动建议输出路径
        if not self.output_path.get() and self.folders:
            base_dir = self.folders[0]
            suggested_name = os.path.basename(base_dir) + "_词频统计.xlsx"
            suggested_path = os.path.join(os.path.dirname(base_dir), suggested_name)
            self.output_path.set(suggested_path)

        # 显示子目录统计
        dir_info_lines = sorted(dir_counts.items())
        if len(dir_info_lines) > 15:
            dir_info = "\n".join(f"  {d}: {n} 个文件" for d, n in dir_info_lines[:12])
            dir_info += f"\n  … 共 {len(dir_info_lines)} 个目录"
        else:
            dir_info = "\n".join(f"  {d}: {n} 个文件" for d, n in dir_info_lines)

        err_info = ""
        if errors:
            err_info = f"\n\n扫描警告（{len(errors)} 个错误）：\n" + "\n".join(errors[:3])

        self._update_status_bar()

        # 自动配置提示
        auto_info = ""
        if best_stkcd:
            auto_info += f"\n  公司代码列 → {best_stkcd}"
        if best_year:
            auto_info += f"\n  日期/年份列 → {best_year}"
        if auto_text_indices:
            auto_text_names = [self._col_display_map[list(self._col_display_map.keys())[i]]
                               for i in auto_text_indices]
            auto_info += f"\n  文本列 → {', '.join(auto_text_names)}"
        if auto_info:
            auto_info = f"\n\n自动检测的列配置：{auto_info}\n（请核实是否正确）"

        messagebox.showinfo(
            "扫描完成",
            f"共扫描到 {len(self.scanned_files)} 个数据文件，\n"
            f"分布在 {len(dir_counts)} 个目录中：\n{dir_info}\n\n"
            f"发现 {len(self.all_columns)} 个数据列。{auto_info}{err_info}"
        )

    def _on_file_select(self, _event=None):
        sel = self.file_listbox.curselection()
        if not sel:
            return
        filepath = self.scanned_files[sel[0]]
        try:
            df = read_data_file(filepath, nrows=100)
        except Exception as e:
            messagebox.showerror("预览失败", str(e))
            return

        self.preview_tree.delete(*self.preview_tree.get_children())
        cols = list(df.columns)
        self.preview_tree["columns"] = cols
        for c in cols:
            self.preview_tree.heading(c, text=c)
            self.preview_tree.column(c, width=130, minwidth=60, stretch=False)
        for _, row in df.iterrows():
            vals = [str(v)[:200] if pd.notna(v) else "" for v in row]
            self.preview_tree.insert("", "end", values=vals)

    # ================================================================
    #  词典管理 — 事件
    # ================================================================

    def _update_dict_info(self):
        n_cat = len(self.dict_mgr.categories())
        n_word = self.dict_mgr.total_word_count()
        self.dict_info_var.set(f"当前词典：{n_cat} 个分类，{n_word} 个关键词")
        self._update_status_bar()

    def _update_status_bar(self):
        n_word = self.dict_mgr.total_word_count()
        n_files = len(self.scanned_files)
        self._status_var.set(f"就绪  |  词典：{n_word} 词  |  数据：{n_files} 文件")

    def _refresh_cat_list(self):
        self.cat_listbox.delete(0, tk.END)
        for c in self.dict_mgr.categories():
            n = len(self.dict_mgr.words(c))
            self.cat_listbox.insert(tk.END, f"{c}  ({n})")
        self._update_dict_info()

    def _refresh_word_list(self, category: str):
        query = self._word_search_var.get().strip().lower() if hasattr(self, "_word_search_var") else ""
        self.word_listbox.delete(0, tk.END)
        for w in self.dict_mgr.words(category):
            if not query or query in w.lower():
                self.word_listbox.insert(tk.END, w)

    def _selected_category(self) -> str | None:
        sel = self.cat_listbox.curselection()
        if not sel:
            return None
        cats = self.dict_mgr.categories()
        return cats[sel[0]] if sel[0] < len(cats) else None

    def _on_cat_select(self, _event=None):
        cat = self._selected_category()
        if cat:
            self._refresh_word_list(cat)

    def _cat_add(self):
        name = simpledialog.askstring("添加分类", "请输入分类名称：", parent=self)
        if name and name.strip():
            self.dict_mgr.add_category(name.strip())
            self._refresh_cat_list()

    def _cat_remove(self):
        cat = self._selected_category()
        if not cat:
            messagebox.showwarning("提示", "请先选择一个分类。")
            return
        if messagebox.askyesno("确认", f"确定删除分类「{cat}」及其所有关键词？"):
            self.dict_mgr.remove_category(cat)
            self.word_listbox.delete(0, tk.END)
            self._refresh_cat_list()

    def _word_add(self):
        cat = self._selected_category()
        if not cat:
            messagebox.showwarning("提示", "请先选择一个分类。")
            return
        word = simpledialog.askstring("添加关键词", f"分类「{cat}」— 请输入关键词：", parent=self)
        if word and word.strip():
            self.dict_mgr.add_word(cat, word.strip())
            self._refresh_word_list(cat)
            self._update_dict_info()

    def _word_batch_add(self):
        cat = self._selected_category()
        if not cat:
            messagebox.showwarning("提示", "请先选择一个分类。")
            return
        dlg = BatchAddDialog(self, title=f"批量添加关键词 — {cat}")
        self.wait_window(dlg)
        if dlg.result:
            for w in dlg.result:
                self.dict_mgr.add_word(cat, w)
            self._refresh_word_list(cat)
            self._refresh_cat_list()

    def _word_remove(self):
        cat = self._selected_category()
        if not cat:
            messagebox.showwarning("提示", "请先选择一个分类。")
            return
        sel = self.word_listbox.curselection()
        if not sel:
            messagebox.showwarning("提示", "请先选择要删除的关键词（可多选）。")
            return
        words = [self.word_listbox.get(i) for i in sel]
        if len(words) > 1:
            if not messagebox.askyesno("确认", f"确定删除选中的 {len(words)} 个关键词？"):
                return
        for word in words:
            self.dict_mgr.remove_word(cat, word)
        self._refresh_word_list(cat)
        self._refresh_cat_list()

    def _cat_rename(self):
        cat = self._selected_category()
        if not cat:
            messagebox.showwarning("提示", "请先选择要重命名的分类。")
            return
        new_name = simpledialog.askstring("重命名分类", f"当前名称：{cat}\n新名称：", parent=self)
        if not new_name or not new_name.strip():
            return
        new_name = new_name.strip()
        if new_name == cat:
            return
        if new_name in self.dict_mgr.data:
            messagebox.showwarning("提示", f"分类「{new_name}」已存在。")
            return
        # Rename: copy data under new key
        words = self.dict_mgr.data.pop(cat)
        self.dict_mgr.data[new_name] = words
        self._refresh_cat_list()
        self.word_listbox.delete(0, tk.END)

    def _word_clear_all(self):
        cat = self._selected_category()
        if not cat:
            messagebox.showwarning("提示", "请先选择一个分类。")
            return
        n = len(self.dict_mgr.words(cat))
        if n == 0:
            return
        if messagebox.askyesno("确认", f"确定清空分类「{cat}」的全部 {n} 个关键词？"):
            self.dict_mgr.data[cat] = []
            self._refresh_word_list(cat)
            self._refresh_cat_list()

    def _on_word_search(self, *_args):
        """根据搜索框内容过滤关键词列表。"""
        cat = self._selected_category()
        if not cat:
            return
        query = self._word_search_var.get().strip().lower()
        self.word_listbox.delete(0, tk.END)
        for w in self.dict_mgr.words(cat):
            if not query or query in w.lower():
                self.word_listbox.insert(tk.END, w)

    def _dict_new(self):
        if self.dict_mgr.total_word_count() > 0:
            if not messagebox.askyesno("确认", "新建词典将清空当前所有分类和关键词，确定？"):
                return
        self.dict_mgr.clear()
        self.cat_listbox.delete(0, tk.END)
        self.word_listbox.delete(0, tk.END)
        self._update_dict_info()

    def _dict_import(self):
        path = filedialog.askopenfilename(
            title="导入词典（Excel 或 TXT）",
            filetypes=[
                ("词典文件", "*.xlsx *.xls *.txt"),
                ("Excel 文件", "*.xlsx *.xls"),
                ("文本文件", "*.txt"),
                ("所有文件", "*.*"),
            ],
        )
        if not path:
            return
        try:
            self.dict_mgr.import_file(path)
            self._refresh_cat_list()
            self.word_listbox.delete(0, tk.END)
            summary = self.dict_mgr.summary_text()
            messagebox.showinfo(
                "导入成功",
                f"共 {len(self.dict_mgr.categories())} 个分类，"
                f"{self.dict_mgr.total_word_count()} 个关键词：\n\n{summary}"
            )
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    def _dict_export(self):
        if self.dict_mgr.total_word_count() == 0:
            messagebox.showwarning("提示", "词典为空，无法导出。")
            return
        path = filedialog.asksaveasfilename(
            title="导出词典 Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if path:
            try:
                self.dict_mgr.export_excel(path)
                messagebox.showinfo("导出成功", f"词典已导出至：{path}")
            except Exception as e:
                messagebox.showerror("导出失败", str(e))

    # ================================================================
    #  设置 — 事件
    # ================================================================

    def _choose_stopwords(self):
        path = filedialog.askopenfilename(
            title="选择停用词文件",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
        )
        if path:
            self.stopwords_path.set(path)

    def _choose_jieba_dict(self):
        path = filedialog.askopenfilename(
            title="选择 jieba 用户词典",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
        )
        if path:
            self.jieba_dict_path.set(path)

    def _toggle_api_key_visibility(self):
        if self._llm_api_key_entry is None:
            return
        self._llm_api_key_entry.configure(
            show="" if self.llm_show_key_var.get() else "*"
        )

    def _on_llm_no_limit_toggle(self):
        """勾选「不限制」时禁用句子上限输入框。"""
        if self._llm_sent_entry is None:
            return
        self._llm_sent_entry.configure(
            state="disabled" if self.var_llm_no_limit.get() else "normal"
        )

    def _on_llm_cache_toggle(self):
        state = "normal" if self.llm_cache_custom_var.get() else "disabled"
        if self._llm_cache_entry:
            self._llm_cache_entry.configure(state=state)
        if self._llm_cache_btn:
            self._llm_cache_btn.configure(state=state)

    def _reset_llm_prompt(self):
        """将自定义提示词文本框重置为内置默认模板，方便用户查看并在此基础上修改。"""
        if self._llm_prompt_text is None:
            return
        try:
            from llm_sentence_analyzer import QwenSentenceAnalyzer
            default = QwenSentenceAnalyzer.SYSTEM_PROMPT
        except Exception:
            default = "（无法读取默认提示词，请检查 llm_sentence_analyzer.py 是否存在）"
        self._llm_prompt_text.delete("1.0", "end")
        self._llm_prompt_text.insert("1.0", default)

    def _choose_llm_cache_path(self):
        path = filedialog.asksaveasfilename(
            title="选择 LLM 缓存文件保存位置",
            defaultextension=".json",
            filetypes=[("JSON 文件", "*.json"), ("所有文件", "*.*")],
        )
        if path:
            self.llm_cache_path_var.set(path)

    def _test_llm_connection(self):
        api_key = self.llm_api_key_var.get().strip()
        base_url = self.llm_base_url_var.get().strip() or "https://dashscope.aliyuncs.com/compatible-mode/v1"
        model = self.llm_model_var.get().strip() or "qwen-plus"
        if not api_key:
            messagebox.showwarning("提示", "请先填写 API Key。")
            return

        def _do_test():
            try:
                from openai import OpenAI
                client = OpenAI(api_key=api_key, base_url=base_url.rstrip("/"), timeout=20.0)
                resp = client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": "请回复 OK"}],
                    max_tokens=10,
                    temperature=0,
                )
                reply = (resp.choices[0].message.content or "").strip()
                self.after(0, lambda: messagebox.showinfo(
                    "连接成功",
                    f"模型：{model}\n接口回复：{reply}\n\nAPI Key 和 Base URL 验证通过。"
                ))
            except ImportError:
                self.after(0, lambda: messagebox.showerror(
                    "缺少依赖", "请先安装 openai 库：\npip install openai"
                ))
            except Exception as exc:
                err = str(exc)
                self.after(0, lambda: messagebox.showerror(
                    "连接失败",
                    f"错误信息：\n{err}\n\n请检查 API Key、Base URL 和网络连接。"
                ))

        threading.Thread(target=_do_test, daemon=True).start()
        messagebox.showinfo("测试中", f"正在连接 {model}，请稍候（最多等待 20 秒）...")

    def _choose_output(self):
        path = filedialog.asksaveasfilename(
            title="选择输出文件路径",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if path:
            self.output_path.set(path)

    # ================================================================
    #  日志 / 进度（线程安全）
    # ================================================================

    def _log(self, msg: str):
        def _do():
            self.log_text.configure(state="normal")
            # v3.0: 日志着色 — 错误红色，完成绿色，警告橙色
            tag = None
            if "错误" in msg or "跳过" in msg or "失败" in msg:
                tag = "error"
            elif "完成" in msg or "成功" in msg:
                tag = "success"
            elif "警告" in msg or "提示" in msg:
                tag = "warn"
            if tag:
                self.log_text.insert(tk.END, msg + "\n", tag)
            else:
                self.log_text.insert(tk.END, msg + "\n")
            self.log_text.see(tk.END)
            self.log_text.configure(state="disabled")
        self.after(0, _do)

    def _update_progress(self, current: int, total: int):
        def _do():
            self.progress["maximum"] = total
            self.progress["value"] = current
        self.after(0, _do)

    def _set_current_file(self, msg: str):
        self.after(0, lambda: self.current_file_var.set(msg))

    # ================================================================
    #  运行分析
    # ================================================================

    def _cancel_analysis(self):
        self._cancel_event.set()
        self._log("正在取消…")
        self.btn_cancel.configure(state="disabled")

    def _start_analysis(self):
        if not self.scanned_files:
            messagebox.showwarning("提示", "请先在「数据选择」扫描文件。")
            return

        col_stkcd = self.combo_stkcd.get()
        col_year = self.combo_year.get()
        if not col_stkcd:
            messagebox.showwarning("提示", "请在「数据选择」配置 公司代码列。")
            return
        if not col_year:
            messagebox.showwarning("提示", "请在「数据选择」配置 日期/年份列。")
            return

        # v3.0: 将显示标签还原为真实列名
        col_map = getattr(self, "_col_display_map", {})
        selected_text_cols = []
        for i in self.col_listbox.curselection():
            label = self.col_listbox.get(i)
            real_col = col_map.get(label, label)
            selected_text_cols.append(real_col)
        if not selected_text_cols:
            messagebox.showwarning("提示", "请在「数据选择」选择至少一个文本列。")
            return

        if self.dict_mgr.total_word_count() == 0:
            messagebox.showwarning("提示", "词典为空，请先在「词典管理」添加或导入词典。")
            return

        if not self.output_path.get():
            messagebox.showwarning("提示", "请在「分析设置」选择输出路径。")
            return

        # 在锁定界面前完成所有参数校验，否则输入错误会让按钮停留在禁用状态。
        try:
            analysis_workers_val = max(1, int(self.analysis_workers_var.get().strip() or "1"))
        except ValueError:
            messagebox.showwarning("提示", "并发进程数必须是正整数。")
            return

        llm_max_sentences = 500
        llm_max_workers = 4
        llm_max_retries = 2
        llm_cache_path = None
        llm_api_key = self.llm_api_key_var.get().strip()
        llm_model = self.llm_model_var.get().strip() or "qwen-plus"
        llm_base_url = self.llm_base_url_var.get().strip() or "https://dashscope.aliyuncs.com/compatible-mode/v1"
        if self.var_llm.get():
            if not self.var_sentences.get():
                if not messagebox.askyesno(
                    "提示",
                    "LLM 句子分析依赖导出命中句子。是否自动启用后继续？",
                ):
                    return
                self.var_sentences.set(True)
            if not llm_api_key and not os.getenv("DASHSCOPE_API_KEY", "").strip():
                messagebox.showwarning(
                    "提示",
                    "启用 LLM 句子分析时需要 API Key。\n请在界面中填写，或先设置环境变量 DASHSCOPE_API_KEY。",
                )
                return
            try:
                if self.var_llm_no_limit.get():
                    llm_max_sentences = 0  # 0 = 无上限
                else:
                    llm_max_sentences = max(1, int(self.llm_max_sentences_var.get().strip() or "500"))
                llm_max_workers = max(1, int(self.llm_max_workers_var.get().strip() or "4"))
                llm_max_retries = max(1, int(self.llm_max_retries_var.get().strip() or "2"))
            except ValueError:
                messagebox.showwarning("提示", "LLM 的数值参数（并发线程、最大重试）必须是正整数。")
                return
            if self.llm_cache_custom_var.get():
                custom = self.llm_cache_path_var.get().strip()
                if custom:
                    llm_cache_path = custom
                else:
                    messagebox.showwarning("提示", "已勾选自定义缓存路径，但未选择文件位置。")
                    return

        # 停用词
        stopwords: set[str] = set()
        if self.var_stopwords.get():
            stopwords = set(DEFAULT_STOPWORDS)
            if self.stopwords_path.get() and os.path.isfile(self.stopwords_path.get()):
                stopwords |= load_stopwords_file(self.stopwords_path.get())

        # 锁定 UI
        self._cancel_event.clear()
        self.btn_start.configure(state="disabled")
        self.btn_cancel.configure(state="normal")
        self.progress["value"] = 0
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state="disabled")
        self.current_file_var.set("准备中…")

        def log_with_status(msg):
            self._log(msg)
            self._set_current_file(msg)

        kwargs = dict(
            files=list(self.scanned_files),
            dict_mgr=self.dict_mgr,
            col_stkcd=col_stkcd,
            col_year=col_year,
            text_columns=selected_text_cols,
            output_path=self.output_path.get(),
            use_regex=self.var_regex.get(),
            use_stopwords=self.var_stopwords.get(),
            stopwords=stopwords,
            use_tf=self.var_tf.get(),
            export_sentences=self.var_sentences.get(),
            preserve_rows=self.var_preserve_rows.get(),
            analyze_llm=self.var_llm.get(),
            llm_api_key=llm_api_key,
            llm_model=llm_model,
            llm_base_url=llm_base_url,
            llm_max_sentences=llm_max_sentences,
            llm_max_workers=llm_max_workers,
            llm_max_retries=llm_max_retries,
            llm_cache_path=llm_cache_path,
            llm_system_prompt=(
                self._llm_prompt_text.get("1.0", "end-1c").strip()
                if self._llm_prompt_text else ""
            ),
            analysis_workers=analysis_workers_val,
            jieba_userdict=self.jieba_dict_path.get(),
            progress_cb=self._update_progress,
            log_cb=log_with_status,
            cancel_event=self._cancel_event,
        )

        thread = threading.Thread(target=self._run_thread, kwargs=kwargs, daemon=True)
        thread.start()

    def _run_thread(self, **kwargs):
        try:
            run_analysis(**kwargs)
            self.after(0, lambda: messagebox.showinfo("完成", "面板数据词频统计已完成！"))
        except Exception as e:
            msg = str(e)
            self._log(f"错误：{msg}")
            if "取消" not in msg:
                self.after(0, lambda: messagebox.showerror("运行出错", msg))
        finally:
            self.after(0, lambda: self.btn_start.configure(state="normal"))
            self.after(0, lambda: self.btn_cancel.configure(state="disabled"))
            self._set_current_file("就绪")


# ============================================================
# 入口
# ============================================================

if __name__ == "__main__":
    # Windows PyInstaller multiprocessing 兼容
    import multiprocessing
    multiprocessing.freeze_support()

    if not _HAS_TKINTER:
        print("错误：当前环境没有安装 tkinter（图形界面库）。")
        print("服务器/无头环境请使用命令行入口：python3 run_server.py --help")
        sys.exit(1)

    # PyInstaller 打包兼容
    if getattr(sys, "frozen", False):
        os.chdir(os.path.dirname(sys.executable))

    app = WordFreqApp()
    app.mainloop()
